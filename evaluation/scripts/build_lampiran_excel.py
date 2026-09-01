"""Extends evaluation/reports/2026-07-23/Lampiran_Evaluasi_ACIF.xlsx with the 2026-07-24
N-gate ablation study results (410 new rows on sheets 2/3) plus two new sheets (Ablation Matrix
Summary, Gate Impact Ranking) with embedded matplotlib charts.

Run from the campus-va/ directory:
    python evaluation/scripts/build_lampiran_excel.py

Reads: evaluation/reports/2026-07-23/Lampiran_Evaluasi_ACIF.xlsx (base to extend)
       evaluation/reports/2026-07-24/raw_exports/*.csv
       evaluation/reports/2026-07-24/ablation_matrix_summary_fixed.csv
       evaluation/reports/2026-07-24/gate_impact_ranking_fixed.csv
       evaluation/reports/2026-07-24/figures/*.png
       backend/app/evaluation/gold_qa_dataset.jsonl
Writes: evaluation/reports/2026-07-24/Lampiran_Evaluasi_ACIF.xlsx
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from eval_data_loader import build_sheet2_rows, build_sheet3_rows, load_gold_qa, load_raw_exports  # noqa: E402

ROOT = Path(__file__).parent.parent.parent  # campus-va/
BASE_XLSX = ROOT / "evaluation/reports/2026-07-23/Lampiran_Evaluasi_ACIF.xlsx"
RAW_DIR = ROOT / "evaluation/reports/2026-07-24/raw_exports"
FIGURES_DIR = ROOT / "evaluation/reports/2026-07-24/figures"
GOLD_QA_PATH = ROOT / "backend/app/evaluation/gold_qa_dataset.jsonl"
OUT_XLSX = ROOT / "evaluation/reports/2026-07-24/Lampiran_Evaluasi_ACIF.xlsx"

HEADER_FILL = PatternFill(start_color="FF1F3B3C", end_color="FF1F3B3C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

ABLATION_CONFIG_NAMES = [
    "gates_all", "gates_none", "gates_1", "gates_1_2", "gates_1_2_3", "gates_1_2_3_4",
    "gates_all_minus_1", "gates_all_minus_2", "gates_all_minus_3", "gates_all_minus_4",
]


def append_rows(ws, rows: list[dict], headers: list[str]) -> None:
    start_row = ws.max_row + 1
    for r_idx, row in enumerate(rows):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=start_row + r_idx, column=c_idx, value=row.get(h))


def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def add_csv_sheet(wb, sheet_name: str, csv_path: Path) -> None:
    ws = wb.create_sheet(sheet_name)
    with open(csv_path, newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append(row)
    style_header(ws, ws.max_column)
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        maxlen = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells[:50])
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 45)


def embed_image(ws, png_path: Path, anchor_cell: str, width_px: int = 700) -> None:
    if not png_path.exists():
        return
    img = XLImage(str(png_path))
    scale = width_px / img.width
    img.width = width_px
    img.height = int(img.height * scale)
    ws.add_image(img, anchor_cell)


def main() -> None:
    print(f"Loading base workbook: {BASE_XLSX}")
    wb = openpyxl.load_workbook(BASE_XLSX)

    dfs = load_raw_exports(RAW_DIR)
    gold_qa = load_gold_qa(GOLD_QA_PATH)

    # --- Sheet 0: Ringkasan -- append 2026-07-24 notes ---
    ws0 = wb["0. Ringkasan"]
    ws0.append([None])
    ws0.append(["=" * 60])
    ws0.append(["PEMBARUAN 2026-07-24: N-Gate ACIF Ablation Study"])
    ws0.append(["Sumber data: VPS Produksi (root@<PRODUCTION_VPS_IP>, database assistant_db)"])
    ws0.append(["10 konfigurasi gate ACIF dijalankan pada tanggal yang sama (2026-07-24), masing-masing"])
    ws0.append(["41 soal gold QA (410 baris baru total) -- lihat sheet '4. Ablation Matrix Summary' dan"])
    ws0.append(["'5. Gate Impact Ranking'. LLM-judge (skor Faithfulness/Relevansi Jawaban) diaktifkan untuk"])
    ws0.append(["pertama kalinya pada run ini -- baris-baris lama (sebelum 2026-07-24) tetap menampilkan"])
    ws0.append(["'N/A - LLM-judge tidak aktif' karena judge memang belum aktif saat run tsb dijalankan."])
    ws0.append([None])
    ws0.append(["CATATAN PENTING -- KETERBANDINGAN: kondisi gates_all/gates_none pada 2026-07-24 TIDAK"])
    ws0.append(["dapat dibandingkan secara numerik langsung dengan angka with_acif/without_acif yang dikutip"])
    ws0.append(["pada sidang 2026-07-20 -- kode retrieval/prompt-boundary/graph-reasoning berubah di antara"])
    ws0.append(["kedua tanggal tsb (ukuran chunk, query rewriter, graph reasoning agent). Data 2026-07-24"])
    ws0.append(["valid untuk analisis relatif antar-gate PADA HARI YANG SAMA (which gate matters), bukan"])
    ws0.append(["sebagai revisi angka yang sudah dikutip pada sidang."])
    ws0.append([None])
    ws0.append(["CATATAN -- precision@3/recall@3: bernilai 0.0 di SEMUA 10 kondisi (bukan hanya sebagian)."])
    ws0.append(["Ini adalah masalah lama (sudah terdeteksi sejak evaluasi 2026-07-15) pada pin"])
    ws0.append(["expected_chunk_ids di gold_qa_dataset.jsonl yang tidak lagi cocok dengan document_chunks.id"])
    ws0.append(["saat ini (kemungkinan karena re-ingestion/re-indexing dokumen). BUKAN bug baru dari ablation"])
    ws0.append(["ini, dan belum diperbaiki pada sesi ini (di luar cakupan tugas evaluasi ini)."])

    # --- Sheet 2: Hasil Evaluasi per Soal -- append 410 rows ---
    ws2 = wb["2. Hasil Evaluasi per Soal"]
    headers2 = [c.value for c in ws2[1]]
    rows2 = build_sheet2_rows(dfs, gold_qa, ABLATION_CONFIG_NAMES)
    append_rows(ws2, rows2, headers2)
    print(f"Sheet 2: +{len(rows2)} rows (total {ws2.max_row - 1})")

    # --- Sheet 3: Log Chat Lengkap -- append 410 rows ---
    ws3 = wb["3. Log Chat Lengkap"]
    headers3 = [c.value for c in ws3[1]]
    rows3 = build_sheet3_rows(dfs)
    start_row = ws3.max_row + 1
    for r_idx, row in enumerate(rows3):
        for c_idx, h in enumerate(headers3, start=1):
            cell = ws3.cell(row=start_row + r_idx, column=c_idx, value=row.get(h))
            cell.alignment = WRAP
    print(f"Sheet 3: +{len(rows3)} rows (total {ws3.max_row - 1})")

    # --- New Sheet 4: Ablation Matrix Summary ---
    add_csv_sheet(wb, "4. Ablation Matrix Summary", ROOT / "evaluation/reports/2026-07-24/ablation_matrix_summary_fixed.csv")
    ws4 = wb["4. Ablation Matrix Summary"]
    anchor_row = ws4.max_row + 3
    embed_image(ws4, FIGURES_DIR / "fig_ablation_quality_per_config.png", f"A{anchor_row}")
    embed_image(ws4, FIGURES_DIR / "fig_ablation_faithfulness_relevance_per_config.png", f"A{anchor_row + 32}")
    embed_image(ws4, FIGURES_DIR / "fig_ablation_latency_per_config.png", f"A{anchor_row + 64}")

    # --- New Sheet 5: Gate Impact Ranking ---
    add_csv_sheet(wb, "5. Gate Impact Ranking", ROOT / "evaluation/reports/2026-07-24/gate_impact_ranking_fixed.csv")
    ws5 = wb["5. Gate Impact Ranking"]
    anchor_row = ws5.max_row + 3
    embed_image(ws5, FIGURES_DIR / "fig_gate_impact_ranking.png", f"A{anchor_row}")

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
