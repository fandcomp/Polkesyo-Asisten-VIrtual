"""Extends evaluation/reports/2026-07-23/Log_Chat_Lengkap.xlsx with the 410 new chat-turn rows
from the 2026-07-24 N-gate ablation study, reusing the same row-builder as
build_lampiran_excel.py's sheet 3 (identical narrative content, this file's schema just adds a
leading 'No' column and keeps 'Session ID').

Run from the campus-va/ directory:
    python evaluation/scripts/build_chatlog_excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

sys.path.insert(0, str(Path(__file__).parent))
from eval_data_loader import build_sheet3_rows, load_raw_exports  # noqa: E402

ROOT = Path(__file__).parent.parent.parent  # campus-va/
BASE_XLSX = ROOT / "evaluation/reports/2026-07-23/Log_Chat_Lengkap.xlsx"
RAW_DIR = ROOT / "evaluation/reports/2026-07-24/raw_exports"
OUT_XLSX = ROOT / "evaluation/reports/2026-07-24/Log_Chat_Lengkap.xlsx"

WRAP = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    print(f"Loading base workbook: {BASE_XLSX}")
    wb = openpyxl.load_workbook(BASE_XLSX)

    dfs = load_raw_exports(RAW_DIR)

    ws0 = wb["0. Ringkasan"]
    ws0.append([None])
    ws0.append(["PEMBARUAN 2026-07-24: +410 chat dari N-Gate ACIF Ablation Study (10 konfigurasi x 41"])
    ws0.append(["soal gold QA). Lihat evaluation/reports/2026-07-24/Lampiran_Evaluasi_ACIF.xlsx sheet"])
    ws0.append(["'4. Ablation Matrix Summary' untuk ringkasan skor per konfigurasi."])

    ws1 = wb["1. Log Chat"]
    headers = [c.value for c in ws1[1]]
    rows = build_sheet3_rows(dfs)

    start_row = ws1.max_row + 1
    existing_no = ws1.cell(row=ws1.max_row, column=1).value
    next_no = int(existing_no) + 1 if isinstance(existing_no, (int, float)) else start_row - 1

    for r_idx, row in enumerate(rows):
        row_with_no = {"No": next_no + r_idx, **row}
        for c_idx, h in enumerate(headers, start=1):
            cell = ws1.cell(row=start_row + r_idx, column=c_idx, value=row_with_no.get(h))
            cell.alignment = WRAP

    print(f"Sheet '1. Log Chat': +{len(rows)} rows (total {ws1.max_row - 1})")

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
