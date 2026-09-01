"""Build the master research dataset workbook for the perception-vs-performance paper.

Restructures the already-validated 2026-08-02 ASQ/scenario linkage data
(evaluation/reports/2026-08-02/raw_joined_by_scenario.csv) into an 11-sheet
research-dataset workbook (00_README ... 10_DATA_QUALITY) suitable for direct
reuse in Python/R/SPSS and for citation as the paper's master dataset.

This script does NOT re-derive statistics from scratch: the six statistical
functions below (descriptive, cronbach_alpha, compute_task_success,
compute_completion_time, compute_latency, compute_correlations,
compute_wilcoxon) are ported verbatim from
build_asq_scenario_excel_2026-08-02.py so that this workbook's own
independent recomputation is a real reconciliation check against that
already-published report, not a tautology.

No new data is collected and no VPS/database access is required -- the sole
input is the already-anonymized, privacy-cleared CSV (R01-R22 codes only, no
trace_id/session_id/real participant codes).

Usage:
    python build_research_dataset_excel_2026-08-10.py --input-csv <path>
"""
from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from scipy import stats

ROOT = Path(__file__).parent.parent.parent  # campus-va/
DEFAULT_INPUT = ROOT / "evaluation" / "reports" / "2026-08-02" / "raw_joined_by_scenario.csv"
OUT_DIR = ROOT / "evaluation" / "reports" / "2026-08-10"

HEADER_FILL = PatternFill(start_color="FF1F3B3C", end_color="FF1F3B3C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13)
WRAP = Alignment(wrap_text=True, vertical="top")

SUCCESS_STATUSES = {"answered", "verified"}
COMPLETION_TIME_OUTLIER_THRESHOLD_S = 3600

REPORTED_VALUES = {
    "ASQ overall mean (n=43)": ("5.76", "reported by user as already-published in the paper"),
    "n observations": ("43", "reported by user as already-published in the paper"),
    "n participants": ("22", "reported by user as already-published in the paper"),
}


# ---------------------------------------------------------------------------
# Statistical functions -- ported verbatim from
# build_asq_scenario_excel_2026-08-02.py so this workbook's recomputation is
# a genuine reconciliation check, not a tautology. Do not "improve" these
# without also re-verifying against that report's published numbers.
# ---------------------------------------------------------------------------

def descriptive(series: pd.Series) -> dict:
    s = series.dropna()
    n = len(s)
    mean = s.mean()
    sd = s.std(ddof=1) if n > 1 else np.nan
    sem = sd / np.sqrt(n) if n > 1 else np.nan
    tcrit = stats.t.ppf(0.975, df=n - 1) if n > 1 else np.nan
    ci_lo, ci_hi = (mean - tcrit * sem, mean + tcrit * sem) if n > 1 else (np.nan, np.nan)
    q1, q3 = (s.quantile(0.25), s.quantile(0.75)) if n > 0 else (np.nan, np.nan)
    return {
        "n": n, "mean": mean, "median": s.median() if n else np.nan, "sd": sd,
        "iqr": q3 - q1 if n else np.nan, "q1": q1, "q3": q3,
        "ci95_lo": ci_lo, "ci95_hi": ci_hi,
    }


def cronbach_alpha(item_df: pd.DataFrame) -> float:
    item_vars = item_df.var(axis=0, ddof=1)
    total_var = item_df.sum(axis=1).var(ddof=1)
    k = item_df.shape[1]
    if total_var == 0:
        return np.nan
    return float((k / (k - 1)) * (1 - item_vars.sum() / total_var))


def compute_task_success(joined: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for scenario in ["S1", "S2"]:
        sub = joined[joined["scenario"] == scenario]
        with_data = sub[sub["has_linked_turn"]]
        n_data = len(with_data)
        n_success = int(with_data["task_success"].sum()) if n_data else 0
        rows.append({
            "scenario": scenario, "n_total_rows": len(sub), "n_with_linked_turn": n_data,
            "n_success": n_success,
            "success_rate": n_success / n_data if n_data else np.nan,
        })
    return pd.DataFrame(rows)


def compute_completion_time(joined: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for scenario in ["S1", "S2"]:
        sub = joined[joined["scenario"] == scenario]
        clean = sub[sub["completion_time_clean"]]
        n_negative = int((~sub["completion_time_valid"]).sum())
        n_outlier = int(sub["completion_time_outlier"].sum())
        desc = descriptive(clean["completion_time_s"])
        rows.append({
            "scenario": scenario, "n_clean": len(clean),
            "n_excluded_negative": n_negative,
            "n_excluded_outlier_gt_1h": n_outlier,
            **desc,
        })
    return pd.DataFrame(rows)


def compute_latency(joined: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for scenario in ["S1", "S2"]:
        sub = joined[joined["scenario"] == scenario]["total_latency_ms"].dropna()
        rows.append({
            "scenario": scenario, "n": len(sub),
            "mean_ms": sub.mean() if len(sub) else np.nan,
            "median_ms": sub.median() if len(sub) else np.nan,
            "p95_ms": sub.quantile(0.95) if len(sub) else np.nan,
            "sd_ms": sub.std(ddof=1) if len(sub) > 1 else np.nan,
        })
    return pd.DataFrame(rows)


def compute_correlations(joined: pd.DataFrame) -> dict:
    out = {}
    sub = joined.dropna(subset=["average_score", "total_latency_ms"])
    pear = stats.pearsonr(sub["average_score"], sub["total_latency_ms"])
    spear = stats.spearmanr(sub["average_score"], sub["total_latency_ms"])
    out["asq_vs_latency"] = {"n": len(sub), "pearson_r": pear.statistic, "pearson_p": pear.pvalue,
                              "spearman_r": spear.statistic, "spearman_p": spear.pvalue}

    sub2 = joined[joined["has_linked_turn"]].dropna(subset=["average_score"])
    pb = stats.pointbiserialr(sub2["task_success"].astype(int), sub2["average_score"])
    out["asq_vs_task_success"] = {"n": len(sub2), "r": pb.correlation, "p": pb.pvalue}

    sub3 = joined[joined["has_linked_turn"]].dropna(subset=["asq_3"])
    pb3 = stats.pointbiserialr(sub3["has_correct_citation"].astype(int), sub3["asq_3"])
    out["asq3_vs_citation"] = {"n": len(sub3), "r": pb3.correlation, "p": pb3.pvalue}
    return out


def compute_wilcoxon(joined: pd.DataFrame) -> dict:
    wide = joined.pivot(index="participant_id", columns="scenario", values="average_score").dropna()
    w, p = stats.wilcoxon(wide["S1"], wide["S2"])
    return {"n_pairs": len(wide), "mean_s1": wide["S1"].mean(), "mean_s2": wide["S2"].mean(),
            "statistic": w, "p": p}


# ---------------------------------------------------------------------------
# Column-spec: single source of truth driving both the real sheets and the
# generated 01_CODEBOOK, so the codebook cannot drift out of sync.
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ColumnSpec:
    sheet: str
    variable_name: str
    data_type: str
    description: str
    unit: str
    allowed_value: str
    missing_rule: str
    analysis_role: str
    source: str
    is_formula: str  # "Y" / "N" / "N/A" (narrative sheets)


CODEBOOK: list[ColumnSpec] = []


def spec(sheet: str, variable_name: str, data_type: str, description: str, unit: str = "",
         allowed_value: str = "", missing_rule: str = "", analysis_role: str = "",
         source: str = "", is_formula: str = "N") -> ColumnSpec:
    c = ColumnSpec(sheet, variable_name, data_type, description, unit, allowed_value,
                    missing_rule, analysis_role, source, is_formula)
    CODEBOOK.append(c)
    return c


# 02_PARTICIPANTS
spec("02_PARTICIPANTS", "participant_id", "string", "Anonymous participant identifier",
     allowed_value="P001-P022", analysis_role="primary key",
     source="derived: R0x (2026-08-02 linkage) -> P0xx by numeric suffix")
spec("02_PARTICIPANTS", "participant_code", "string", "Alias of participant_id; no separate real code retained",
     analysis_role="identifier", source="formula: same as participant_id", is_formula="Y")
spec("02_PARTICIPANTS", "participant_group", "string", "Participant category (student/prospective student/staff)",
     allowed_value="NA", missing_rule="not recorded anywhere in the source data",
     analysis_role="grouping (unavailable)", source="NA")
spec("02_PARTICIPANTS", "completed_s1", "boolean", "Whether this participant has an S1 observation",
     source="COUNTIFS over 03_OBSERVATIONS", is_formula="Y")
spec("02_PARTICIPANTS", "completed_s2", "boolean", "Whether this participant has an S2 observation",
     source="COUNTIFS over 03_OBSERVATIONS", is_formula="Y")
spec("02_PARTICIPANTS", "total_observations", "integer", "Count of observation rows for this participant",
     source="COUNTIF over 03_OBSERVATIONS", is_formula="Y")
spec("02_PARTICIPANTS", "notes", "string", "Free-text note (e.g. missing scenario)", source="derived")

# 03_OBSERVATIONS
spec("03_OBSERVATIONS", "observation_id", "string", "Unique observation identifier",
     allowed_value="OBS001-OBS043", analysis_role="primary key",
     source="derived: sequential, sorted by participant_id then scenario_id")
spec("03_OBSERVATIONS", "participant_id", "string", "Participant this observation belongs to",
     analysis_role="foreign key -> 02_PARTICIPANTS", source="from raw_joined_by_scenario.csv (r0x)")
spec("03_OBSERVATIONS", "scenario_id", "category", "Scenario", allowed_value="S1/S2",
     source="raw_joined_by_scenario.csv:scenario")
spec("03_OBSERVATIONS", "scenario_order", "integer", "1=S1, 2=S2",
     source="formula: IF(scenario_id=\"S1\",1,2)", is_formula="Y")
spec("03_OBSERVATIONS", "session_id", "string", "Production session identifier",
     allowed_value="NA", missing_rule="scrubbed from the safe/anonymized export for participant privacy",
     source="NA")
spec("03_OBSERVATIONS", "timestamp_start", "datetime", "Scenario start timestamp",
     allowed_value="NA", missing_rule="not present in the safe per-observation export", source="NA")
spec("03_OBSERVATIONS", "timestamp_end", "datetime", "Scenario end timestamp",
     allowed_value="NA", missing_rule="not present in the safe per-observation export", source="NA")
spec("03_OBSERVATIONS", "query_text", "string", "User question text",
     allowed_value="NA", missing_rule="not present in the safe per-observation export", source="NA")
spec("03_OBSERVATIONS", "response_id", "string", "System response identifier",
     allowed_value="NA", missing_rule="not present in the safe per-observation export", source="NA")
spec("03_OBSERVATIONS", "task_status", "category", "Production answer_status for this turn",
     allowed_value="verified/out_of_domain/insufficient_context/fallback_enforced/needs_clarification/blank",
     missing_rule="blank when no trace_id was linked (1 row: participant with no matched production turn)",
     source="raw_joined_by_scenario.csv:answer_status")
spec("03_OBSERVATIONS", "data_source", "string", "Provenance of this row", source="derived: source file + row")
spec("03_OBSERVATIONS", "missing_flag", "boolean", "TRUE if task_status is blank",
     source="formula: IF(task_status=\"\",TRUE,FALSE)", is_formula="Y")
spec("03_OBSERVATIONS", "missing_reason", "string", "Why task_status is blank, if applicable", source="derived")

# 04_ASQ_RESPONSES
spec("04_ASQ_RESPONSES", "observation_id", "string", "FK to 03_OBSERVATIONS",
     analysis_role="foreign key", source="formula: ='03_OBSERVATIONS' same row", is_formula="Y")
spec("04_ASQ_RESPONSES", "participant_id", "string", "FK to 02_PARTICIPANTS",
     source="formula: ='03_OBSERVATIONS' same row", is_formula="Y")
spec("04_ASQ_RESPONSES", "scenario_id", "category", "S1/S2", source="formula: ='03_OBSERVATIONS' same row", is_formula="Y")
spec("04_ASQ_RESPONSES", "asq_q1_ease", "integer", "ASQ item 1: ease of completing the task",
     unit="1-7 Likert", allowed_value="1-7", source="raw_joined_by_scenario.csv:asq_1")
spec("04_ASQ_RESPONSES", "asq_q2_time", "integer", "ASQ item 2: satisfaction with time taken",
     unit="1-7 Likert", allowed_value="1-7", source="raw_joined_by_scenario.csv:asq_2")
spec("04_ASQ_RESPONSES", "asq_q3_information_support", "integer",
     "ASQ item 3: satisfaction with information/citations/messages",
     unit="1-7 Likert", allowed_value="1-7", source="raw_joined_by_scenario.csv:asq_3")
spec("04_ASQ_RESPONSES", "asq_mean", "decimal", "Mean of the 3 ASQ items", unit="1-7",
     analysis_role="primary DV (perceived usability)", source="formula: AVERAGE(q1,q2,q3)", is_formula="Y")
spec("04_ASQ_RESPONSES", "asq_complete", "boolean", "TRUE iff all 3 items present",
     source="formula: IF(COUNT(q1:q3)=3,TRUE,FALSE)", is_formula="Y")
spec("04_ASQ_RESPONSES", "asq_notes", "string", "Free-text note", source="derived (source notes column empty)")

# 05_TASK_PERFORMANCE
spec("05_TASK_PERFORMANCE", "observation_id", "string", "FK", source="same row as 03", is_formula="Y")
spec("05_TASK_PERFORMANCE", "participant_id", "string", "FK", source="same row as 03", is_formula="Y")
spec("05_TASK_PERFORMANCE", "scenario_id", "category", "S1/S2", source="same row as 03", is_formula="Y")
spec("05_TASK_PERFORMANCE", "task_success", "binary", "1=success, 0=not success, blank=no linked turn",
     allowed_value="0/1/blank",
     analysis_role="primary objective-performance DV",
     source="raw_joined_by_scenario.csv:task_success (rule: answer_status in {answered, verified})")
spec("05_TASK_PERFORMANCE", "task_status", "category", "same as 03_OBSERVATIONS.task_status",
     source="formula: ='03_OBSERVATIONS' same row", is_formula="Y")
spec("05_TASK_PERFORMANCE", "completion_time_sec", "numeric", "Completion time proxy", unit="seconds",
     source="raw_joined_by_scenario.csv:completion_time_s")
spec("05_TASK_PERFORMANCE", "completion_time_source", "string",
     "Constant: this is a proxy (ASQ-submission timestamp delta), not a stopwatch measurement",
     source="constant")
spec("05_TASK_PERFORMANCE", "success_evidence", "string", "Evidence text derived from task_status", source="derived")
spec("05_TASK_PERFORMANCE", "evaluator_notes", "string",
     "Flags negative/outlier completion-time rows per source's own validity flags", source="derived")

# 06_SYSTEM_METRICS
spec("06_SYSTEM_METRICS", "observation_id", "string", "FK", source="same row as 03", is_formula="Y")
spec("06_SYSTEM_METRICS", "participant_id", "string", "FK", source="same row as 03", is_formula="Y")
spec("06_SYSTEM_METRICS", "scenario_id", "category", "S1/S2", source="same row as 03", is_formula="Y")
spec("06_SYSTEM_METRICS", "latency_ms", "numeric", "Total end-to-end latency for this turn", unit="milliseconds",
     source="raw_joined_by_scenario.csv:total_latency_ms")
for col in ["retrieval_latency_ms", "llm_latency_ms", "retrieved_chunks", "graph_nodes_retrieved",
            "graph_relationships", "token_input", "token_output", "model_name", "request_status"]:
    spec("06_SYSTEM_METRICS", col, "numeric/string", f"{col} (not available at this grain)",
         allowed_value="NA", missing_rule="not present in the safe per-observation export; only total_latency_ms was retained",
         source="NA")
spec("06_SYSTEM_METRICS", "missing_reason", "string", "Summary note for the NA columns above", source="derived")

# 07_CITATION_EVALUATION
spec("07_CITATION_EVALUATION", "observation_id", "string", "FK", source="same row as 03", is_formula="Y")
spec("07_CITATION_EVALUATION", "participant_id", "string", "FK", source="same row as 03", is_formula="Y")
spec("07_CITATION_EVALUATION", "scenario_id", "category", "S1/S2", source="same row as 03", is_formula="Y")
spec("07_CITATION_EVALUATION", "n_citations", "integer", "Number of citations returned in this turn",
     source="raw_joined_by_scenario.csv:n_citations")
spec("07_CITATION_EVALUATION", "n_valid_citations", "integer", "Number of those citations judged valid",
     source="raw_joined_by_scenario.csv:n_valid_citations")
spec("07_CITATION_EVALUATION", "has_correct_citation", "boolean", "TRUE iff n_valid_citations > 0",
     analysis_role="citation-correctness DV (per-observation, human study)",
     source="formula: IF(n_valid_citations=\"\",\"\",n_valid_citations>0)", is_formula="Y")
spec("07_CITATION_EVALUATION", "citation_evaluation_method", "string",
     "Per-observation, joined via trace_id to production citation_evaluation_logs", source="constant")
spec("07_CITATION_EVALUATION", "evaluator_notes", "string",
     "Explicitly NOT the separate 41-question automated gold-QA citation-correct rate (39.0%) -- "
     "that is a different instrument, not tied to these 22 participants; see 09 methodology notes",
     source="constant")

# 08_MERGED_ANALYSIS
spec("08_MERGED_ANALYSIS", "observation_id", "string", "PK / join key", source="same row as 03", is_formula="Y")
spec("08_MERGED_ANALYSIS", "participant_id", "string", "FK", source="INDEX/MATCH into 04 on observation_id", is_formula="Y")
spec("08_MERGED_ANALYSIS", "scenario_id", "category", "S1/S2", source="INDEX/MATCH into 04 on observation_id", is_formula="Y")
spec("08_MERGED_ANALYSIS", "asq_q1_ease", "integer", "same as 04.asq_q1_ease", source="INDEX/MATCH into 04", is_formula="Y")
spec("08_MERGED_ANALYSIS", "asq_q2_time", "integer", "same as 04.asq_q2_time", source="INDEX/MATCH into 04", is_formula="Y")
spec("08_MERGED_ANALYSIS", "asq_q3_information_support", "integer", "same as 04.asq_q3_information_support",
     source="INDEX/MATCH into 04", is_formula="Y")
spec("08_MERGED_ANALYSIS", "asq_mean", "decimal", "same as 04.asq_mean", source="INDEX/MATCH into 04", is_formula="Y")
spec("08_MERGED_ANALYSIS", "task_success", "binary", "same as 05.task_success", source="INDEX/MATCH into 05", is_formula="Y")
spec("08_MERGED_ANALYSIS", "task_status", "category", "same as 05.task_status", source="INDEX/MATCH into 05", is_formula="Y")
spec("08_MERGED_ANALYSIS", "completion_time_sec", "numeric", "same as 05.completion_time_sec",
     source="INDEX/MATCH into 05", is_formula="Y")
spec("08_MERGED_ANALYSIS", "latency_ms", "numeric", "same as 06.latency_ms", source="INDEX/MATCH into 06", is_formula="Y")
spec("08_MERGED_ANALYSIS", "has_correct_citation", "boolean", "same as 07.has_correct_citation",
     source="INDEX/MATCH into 07", is_formula="Y")
spec("08_MERGED_ANALYSIS", "complete_case_asq", "boolean", "TRUE iff asq_mean present",
     source="formula: IF(asq_mean<>\"\",TRUE,FALSE)", is_formula="Y")
spec("08_MERGED_ANALYSIS", "complete_case_performance", "boolean", "TRUE iff task_success present",
     source="formula: IF(task_success<>\"\",TRUE,FALSE)", is_formula="Y")
spec("08_MERGED_ANALYSIS", "paired_scenario_available", "boolean",
     "TRUE iff this participant has both an S1 and S2 row",
     source="formula: COUNTIFS(participant_id range, this participant)=2", is_formula="Y")

# 09/10 narrative
spec("09_STATISTICAL_SUMMARY", "(various)", "mixed",
     "Descriptive stats, Cronbach's alpha, correlations, Wilcoxon test, reconciliation table",
     analysis_role="primary analysis output",
     source="live formulas for descriptive/Cronbach/Pearson/point-biserial; Python-computed values "
            "(scipy, same methods as build_asq_scenario_excel_2026-08-02.py) for Spearman rho/p and "
            "Wilcoxon W/p, since Excel has no native exact Wilcoxon function", is_formula="mixed")
spec("10_DATA_QUALITY", "(various)", "mixed", "Automated data-quality audit checks",
     source="live formulas (COUNTIFS/SUMPRODUCT) over sheets 02-08", is_formula="Y")


# ---------------------------------------------------------------------------
# Data loading + shaping
# ---------------------------------------------------------------------------

def load_and_shape(input_csv: Path) -> pd.DataFrame:
    df = pd.read_csv(input_csv)
    df = df.sort_values(["r0x", "scenario"]).reset_index(drop=True)

    r0x_sorted = sorted(df["r0x"].unique(), key=lambda x: int(x[1:]))
    r0x_to_pid = {r0x: f"P{int(r0x[1:]):03d}" for r0x in r0x_sorted}
    df["participant_id"] = df["r0x"].map(r0x_to_pid)
    df = df.sort_values(["participant_id", "scenario"]).reset_index(drop=True)
    df["observation_id"] = [f"OBS{i + 1:03d}" for i in range(len(df))]

    df["has_linked_turn"] = df["answer_status"].notna()
    return df


# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------

def style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def autosize(ws: Worksheet, max_width: int = 55) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        maxlen = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells[:200])
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), max_width)


def add_table(ws: Worksheet, name: str, ref: str, style: str = "TableStyleMedium2") -> None:
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    ws.add_table(tbl)


def add_list_validation(ws: Worksheet, col_letter: str, first_row: int, last_row: int,
                         choices: list[str]) -> None:
    formula = '"' + ",".join(choices) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def nv(v):
    """Normalize a python/pandas value for openpyxl cell writing."""
    if v is None:
        return None
    if isinstance(v, (np.floating, float)):
        if pd.isna(v):
            return None
        return round(float(v), 6)
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.bool_, bool)):
        return bool(v)
    if isinstance(v, str) and v == "":
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_00_readme(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("00_README")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 100
    r = 1
    ws.cell(row=r, column=1, value="Master Research Dataset -- Perceived Usability vs Objective "
                                    "Task Performance").font = Font(bold=True, size=15)
    r += 2

    def kv(label: str, value: str, wrap: bool = True) -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=value)
        if wrap:
            c.alignment = WRAP
        r += 1

    kv("Title", "Evaluating the Alignment Between Perceived Usability and Objective Task "
                "Performance in a GraphRAG-Based Campus Virtual Assistant")
    kv("Purpose of this dataset", "Master, reusable observation-level dataset joining the human "
                                   "ASQ usability study to objective production task-performance, "
                                   "latency, and citation-correctness records, for direct "
                                   "statistical analysis in Excel, Python, R, or SPSS.")
    kv("Dataset build date", "2026-08-10")
    kv("Data sources", "Sole source: campus-va/evaluation/reports/2026-08-02/"
                        "raw_joined_by_scenario.csv -- an already-anonymized (R01-R22, no PII), "
                        "already-validated join of the 2026-07-16 human ASQ study "
                        "(evaluation/reports/2026-07-16/sus_asq_final/) to production "
                        "chat_evaluation_logs/citation_evaluation_logs via trace_id. See "
                        "evaluation/reports/2026-08-02/README.md for the original build "
                        "methodology and evaluation/scripts/build_asq_scenario_excel_2026-08-02.py "
                        "for the reused statistical functions.")
    kv("Reconstruction method", "This workbook independently re-derives every statistic from the "
                                 "same source CSV using the same computational methods as the "
                                 "already-published 2026-08-02 report (functions ported verbatim, "
                                 "see build_research_dataset_excel_2026-08-10.py), then compares "
                                 "the recomputed values against that report and against the "
                                 "numbers explicitly reported in the paper (sheet "
                                 "09_STATISTICAL_SUMMARY, reconciliation table). No respondent, "
                                 "score, or statistic in this workbook was invented, adjusted, or "
                                 "back-fit to match a target number.")
    kv("n participants", f"{df['participant_id'].nunique()}")
    kv("n observations", f"{len(df)}")
    kv("n scenarios", "2 (S1 = normal campus/registration questions, S2 = difficult/ambiguous/"
                       "manipulative questions)")
    kv("Why 43, not 44 (22x2)", "One participant (P016) has only an S1 observation -- no S2 row "
                                 "exists in the source data. This is a genuine gap in the original "
                                 "study, not an artifact of this reconstruction; see "
                                 "10_DATA_QUALITY and 02_PARTICIPANTS.")
    kv("Meaning of NA", "NA denotes a field that is genuinely unavailable in the traceable source "
                         "data (e.g. scrubbed for privacy, or never captured at this grain). Every "
                         "NA has an accompanying missing_reason. No NA was estimated or imputed.")
    kv("SUS data (out of scope)", "A related instrument, the System Usability Scale (SUS, 21 "
                                   "respondents, mean 69.05/100), was also collected in this "
                                   "project but is out of scope for this ASQ-focused paper and is "
                                   "deliberately excluded from this workbook. See "
                                   "evaluation/reports/2026-07-16/sus_asq_final/ if needed later.")
    kv("Citation-correctness scope warning", "This dataset's citation-correctness figures "
                                              "(07_CITATION_EVALUATION, 08, 09) are per-observation "
                                              "values tied to the 22 real participants. A separate, "
                                              "unrelated 41-question automated gold-QA benchmark "
                                              "also reports a citation-correct rate (39.0%) -- that "
                                              "is a DIFFERENT instrument, not tied to these "
                                              "participants, and is not merged into this dataset's "
                                              "figures anywhere.")
    kv("Formula vs. value policy", "04_ASQ_RESPONSES.asq_mean, all cross-sheet joins in "
                                    "08_MERGED_ANALYSIS, and most of 09/10 use live Excel formulas "
                                    "(AVERAGE/COUNTIFS/VAR.S/CORREL/T.DIST.2T/CONFIDENCE.T/INDEX/"
                                    "MATCH) that recalculate automatically when this file is opened "
                                    "in Excel/LibreOffice/Google Sheets. Two statistics in "
                                    "09_STATISTICAL_SUMMARY -- Spearman rho/p and the Wilcoxon "
                                    "signed-rank test -- are computed once in Python (scipy, same "
                                    "method as the source report) and written as fixed values, "
                                    "because Excel has no native exact Wilcoxon function and a "
                                    "hand-built approximation risks disagreeing with the standard "
                                    "method for reasons unrelated to the data. See 01_CODEBOOK's "
                                    "is_formula column for the authoritative per-variable answer.")
    kv("Primary key / foreign key map",
       "02_PARTICIPANTS.participant_id (PK) <- referenced by 03-08.participant_id (FK). "
       "03_OBSERVATIONS.observation_id (PK) <- referenced by 04-08.observation_id (FK). "
       "03_OBSERVATIONS is the row-order anchor for 04-07 (identical 43-row order); "
       "08_MERGED_ANALYSIS joins via INDEX/MATCH on observation_id, not row-order assumption.")
    kv("Sheets in this workbook",
       "00_README, 01_CODEBOOK, 02_PARTICIPANTS, 03_OBSERVATIONS, 04_ASQ_RESPONSES, "
       "05_TASK_PERFORMANCE, 06_SYSTEM_METRICS, 07_CITATION_EVALUATION, 08_MERGED_ANALYSIS, "
       "09_STATISTICAL_SUMMARY, 10_DATA_QUALITY")
    kv("Data integrity statement",
       "Dataset ini disusun dari data eksperimen, log sistem, dan hasil evaluasi yang tersedia. "
       "Missing values tidak diimputasi kecuali dinyatakan secara eksplisit. No respondent, score, "
       "or statistic was fabricated to match a reported paper value; see the reconciliation table "
       "in 09_STATISTICAL_SUMMARY and dataset_reconstruction_report.md for an explicit comparison.")


def build_01_codebook(wb) -> None:
    ws = wb.create_sheet("01_CODEBOOK")
    headers = ["variable_name", "sheet", "data_type", "description", "unit", "allowed_value",
               "missing_rule", "analysis_role", "source", "is_formula"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    for i, c in enumerate(CODEBOOK, start=2):
        ws.cell(row=i, column=1, value=c.variable_name)
        ws.cell(row=i, column=2, value=c.sheet)
        ws.cell(row=i, column=3, value=c.data_type)
        ws.cell(row=i, column=4, value=c.description)
        ws.cell(row=i, column=5, value=c.unit)
        ws.cell(row=i, column=6, value=c.allowed_value)
        ws.cell(row=i, column=7, value=c.missing_rule)
        ws.cell(row=i, column=8, value=c.analysis_role)
        ws.cell(row=i, column=9, value=c.source)
        ws.cell(row=i, column=10, value=c.is_formula)
    last_row = len(CODEBOOK) + 1
    add_table(ws, "tbl_01_codebook", f"A1:J{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws, max_width=45)


def build_02_participants(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("02_PARTICIPANTS")
    headers = ["participant_id", "participant_code", "participant_group", "completed_s1",
               "completed_s2", "total_observations", "notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    pids = sorted(df["participant_id"].unique(), key=lambda x: int(x[1:]))
    no_s2 = set(df[df["scenario"] == "S1"]["participant_id"]) - set(df[df["scenario"] == "S2"]["participant_id"])
    for i, pid in enumerate(pids, start=2):
        ws.cell(row=i, column=1, value=pid)
        ws.cell(row=i, column=2, value=f"=A{i}")
        ws.cell(row=i, column=3, value="NA")
        ws.cell(row=i, column=4,
                value=f'=COUNTIFS(\'03_OBSERVATIONS\'!$B:$B,A{i},\'03_OBSERVATIONS\'!$C:$C,"S1")>0')
        ws.cell(row=i, column=5,
                value=f'=COUNTIFS(\'03_OBSERVATIONS\'!$B:$B,A{i},\'03_OBSERVATIONS\'!$C:$C,"S2")>0')
        ws.cell(row=i, column=6, value=f"=COUNTIF('03_OBSERVATIONS'!$B:$B,A{i})")
        ws.cell(row=i, column=7,
                value="No S2 observation on record (see 10_DATA_QUALITY)" if pid in no_s2 else "")
    last_row = len(pids) + 1
    add_table(ws, "tbl_02_participants", f"A1:G{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws)


def build_03_observations(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("03_OBSERVATIONS")
    headers = ["observation_id", "participant_id", "scenario_id", "scenario_order", "session_id",
               "timestamp_start", "timestamp_end", "query_text", "response_id", "task_status",
               "data_source", "missing_flag", "missing_reason"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=row["observation_id"])
        ws.cell(row=i, column=2, value=row["participant_id"])
        ws.cell(row=i, column=3, value=row["scenario"])
        ws.cell(row=i, column=4, value=f'=IF(C{i}="S1",1,2)')
        ws.cell(row=i, column=5, value="NA")
        ws.cell(row=i, column=6, value="NA")
        ws.cell(row=i, column=7, value="NA")
        ws.cell(row=i, column=8, value="NA")
        ws.cell(row=i, column=9, value="NA")
        task_status = nv(row["answer_status"])
        ws.cell(row=i, column=10, value=task_status if task_status is not None else "")
        ws.cell(row=i, column=11,
                value=f"raw_joined_by_scenario.csv row {i - 1} (r0x={row['r0x']})")
        ws.cell(row=i, column=12, value=f'=IF(J{i}="",TRUE,FALSE)')
        ws.cell(row=i, column=13,
                value="no linked trace_id for this turn in the production source" if task_status is None else "")
    last_row = len(df) + 1
    add_table(ws, "tbl_03_observations", f"A1:M{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    add_list_validation(ws, "C", 2, last_row, ["S1", "S2"])
    add_list_validation(ws, "J", 2, last_row,
                         ["verified", "out_of_domain", "insufficient_context", "fallback_enforced",
                          "needs_clarification"])
    autosize(ws)


def build_04_asq(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("04_ASQ_RESPONSES")
    headers = ["observation_id", "participant_id", "scenario_id", "asq_q1_ease", "asq_q2_time",
               "asq_q3_information_support", "asq_mean", "asq_complete", "asq_notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=f"='03_OBSERVATIONS'!A{i}")
        ws.cell(row=i, column=2, value=f"='03_OBSERVATIONS'!B{i}")
        ws.cell(row=i, column=3, value=f"='03_OBSERVATIONS'!C{i}")
        ws.cell(row=i, column=4, value=nv(row["asq_1"]))
        ws.cell(row=i, column=5, value=nv(row["asq_2"]))
        ws.cell(row=i, column=6, value=nv(row["asq_3"]))
        ws.cell(row=i, column=7, value=f"=AVERAGE(D{i}:F{i})")
        ws.cell(row=i, column=8, value=f"=IF(COUNT(D{i}:F{i})=3,TRUE,FALSE)")
        ws.cell(row=i, column=9, value="")
    last_row = len(df) + 1
    add_table(ws, "tbl_04_asq", f"A1:I{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws)


def build_05_task_performance(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("05_TASK_PERFORMANCE")
    headers = ["observation_id", "participant_id", "scenario_id", "task_success", "task_status",
               "completion_time_sec", "completion_time_source", "success_evidence",
               "evaluator_notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=f"='03_OBSERVATIONS'!A{i}")
        ws.cell(row=i, column=2, value=f"='03_OBSERVATIONS'!B{i}")
        ws.cell(row=i, column=3, value=f"='03_OBSERVATIONS'!C{i}")
        ts = nv(row["task_success"])
        ws.cell(row=i, column=4, value=ts if ts is not None else "")
        ws.cell(row=i, column=5, value=f"='03_OBSERVATIONS'!J{i}")
        ws.cell(row=i, column=6, value=nv(row["completion_time_s"]))
        ws.cell(row=i, column=7, value="proxy: ASQ-submission timestamp delta, not stopwatch")
        status = nv(row["answer_status"])
        ws.cell(row=i, column=8, value=f"answer_status={status}" if status else "no linked turn")
        notes = []
        if not bool(row["completion_time_valid"]):
            notes.append("negative/invalid completion-time proxy (excluded from clean stats)")
        if bool(row["completion_time_outlier"]):
            notes.append(">1h outlier, excluded from clean completion-time stats")
        ws.cell(row=i, column=9, value="; ".join(notes))
    last_row = len(df) + 1
    add_table(ws, "tbl_05_task_performance", f"A1:I{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws)


def build_06_system_metrics(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("06_SYSTEM_METRICS")
    headers = ["observation_id", "participant_id", "scenario_id", "latency_ms",
               "retrieval_latency_ms", "llm_latency_ms", "retrieved_chunks",
               "graph_nodes_retrieved", "graph_relationships", "token_input", "token_output",
               "model_name", "request_status", "missing_reason"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    na_reason = ("not present in the per-observation safe export; only total end-to-end latency "
                 "was retained from production chat_evaluation_logs")
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=f"='03_OBSERVATIONS'!A{i}")
        ws.cell(row=i, column=2, value=f"='03_OBSERVATIONS'!B{i}")
        ws.cell(row=i, column=3, value=f"='03_OBSERVATIONS'!C{i}")
        ws.cell(row=i, column=4, value=nv(row["total_latency_ms"]))
        for j in range(5, 14):
            ws.cell(row=i, column=j, value="NA")
        ws.cell(row=i, column=14, value=na_reason)
    last_row = len(df) + 1
    add_table(ws, "tbl_06_system_metrics", f"A1:N{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws)


def build_07_citation(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("07_CITATION_EVALUATION")
    headers = ["observation_id", "participant_id", "scenario_id", "n_citations",
               "n_valid_citations", "has_correct_citation", "citation_evaluation_method",
               "evaluator_notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    warn_note = ("Per-observation value tied to these 22 participants via trace_id. NOT the "
                 "separate 41-question automated gold-QA citation-correct rate (39.0%) -- a "
                 "different instrument; see 09_STATISTICAL_SUMMARY methodology notes.")
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=f"='03_OBSERVATIONS'!A{i}")
        ws.cell(row=i, column=2, value=f"='03_OBSERVATIONS'!B{i}")
        ws.cell(row=i, column=3, value=f"='03_OBSERVATIONS'!C{i}")
        ws.cell(row=i, column=4, value=nv(row["n_citations"]))
        ws.cell(row=i, column=5, value=nv(row["n_valid_citations"]))
        ws.cell(row=i, column=6, value=f'=IF(E{i}="","",E{i}>0)')
        ws.cell(row=i, column=7,
                value="joined via trace_id to production citation_evaluation_logs")
        ws.cell(row=i, column=8, value=warn_note if i == 2 else "")
    last_row = len(df) + 1
    add_table(ws, "tbl_07_citation", f"A1:H{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws)


def build_08_merged(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("08_MERGED_ANALYSIS")
    headers = ["observation_id", "participant_id", "scenario_id", "asq_q1_ease", "asq_q2_time",
               "asq_q3_information_support", "asq_mean", "task_success", "task_status",
               "completion_time_sec", "latency_ms", "has_correct_citation", "complete_case_asq",
               "complete_case_performance", "paired_scenario_available"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    n = len(df)
    for i in range(2, n + 2):
        ws.cell(row=i, column=1, value=f"='03_OBSERVATIONS'!A{i}")
        ws.cell(row=i, column=2,
                value=f"=INDEX('04_ASQ_RESPONSES'!$B:$B,MATCH($A{i},'04_ASQ_RESPONSES'!$A:$A,0))")
        ws.cell(row=i, column=3,
                value=f"=INDEX('04_ASQ_RESPONSES'!$C:$C,MATCH($A{i},'04_ASQ_RESPONSES'!$A:$A,0))")
        ws.cell(row=i, column=4,
                value=f"=INDEX('04_ASQ_RESPONSES'!$D:$D,MATCH($A{i},'04_ASQ_RESPONSES'!$A:$A,0))")
        ws.cell(row=i, column=5,
                value=f"=INDEX('04_ASQ_RESPONSES'!$E:$E,MATCH($A{i},'04_ASQ_RESPONSES'!$A:$A,0))")
        ws.cell(row=i, column=6,
                value=f"=INDEX('04_ASQ_RESPONSES'!$F:$F,MATCH($A{i},'04_ASQ_RESPONSES'!$A:$A,0))")
        ws.cell(row=i, column=7,
                value=f"=INDEX('04_ASQ_RESPONSES'!$G:$G,MATCH($A{i},'04_ASQ_RESPONSES'!$A:$A,0))")
        ws.cell(row=i, column=8,
                value=f"=INDEX('05_TASK_PERFORMANCE'!$D:$D,MATCH($A{i},'05_TASK_PERFORMANCE'!$A:$A,0))")
        ws.cell(row=i, column=9,
                value=f"=INDEX('05_TASK_PERFORMANCE'!$E:$E,MATCH($A{i},'05_TASK_PERFORMANCE'!$A:$A,0))")
        ws.cell(row=i, column=10,
                value=f"=INDEX('05_TASK_PERFORMANCE'!$F:$F,MATCH($A{i},'05_TASK_PERFORMANCE'!$A:$A,0))")
        ws.cell(row=i, column=11,
                value=f"=INDEX('06_SYSTEM_METRICS'!$D:$D,MATCH($A{i},'06_SYSTEM_METRICS'!$A:$A,0))")
        ws.cell(row=i, column=12,
                value=f"=INDEX('07_CITATION_EVALUATION'!$F:$F,MATCH($A{i},'07_CITATION_EVALUATION'!$A:$A,0))")
        ws.cell(row=i, column=13, value=f'=IF(G{i}<>"",TRUE,FALSE)')
        ws.cell(row=i, column=14, value=f'=IF(H{i}<>"",TRUE,FALSE)')
        ws.cell(row=i, column=15, value=f"=COUNTIF($B$2:$B${n + 1},B{i})=2")
    last_row = n + 1
    add_table(ws, "tbl_08_merged", f"A1:O{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws)


HELPER_COLS = [
    # (name, formula_template) -- formula_template uses {i} for the row number.
    # All reference '08_MERGED_ANALYSIS' same-row (row i in 09's helper block ==
    # row i in 08_MERGED_ANALYSIS, since both mirror 03_OBSERVATIONS' row order).
    ("asq_mean_s1", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S1",\'08_MERGED_ANALYSIS\'!G{i},"")'),
    ("asq_mean_s2", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S2",\'08_MERGED_ANALYSIS\'!G{i},"")'),
    ("q1_s1", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S1",\'08_MERGED_ANALYSIS\'!D{i},"")'),
    ("q2_s1", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S1",\'08_MERGED_ANALYSIS\'!E{i},"")'),
    ("q3_s1", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S1",\'08_MERGED_ANALYSIS\'!F{i},"")'),
    ("q1_s2", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S2",\'08_MERGED_ANALYSIS\'!D{i},"")'),
    ("q2_s2", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S2",\'08_MERGED_ANALYSIS\'!E{i},"")'),
    ("q3_s2", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S2",\'08_MERGED_ANALYSIS\'!F{i},"")'),
    ("total_s1", '=IF(P{i}="","",P{i}+Q{i}+R{i})'),
    ("total_s2", '=IF(S{i}="","",S{i}+T{i}+U{i})'),
    ("completion_s1_clean",
     '=IF(AND(\'08_MERGED_ANALYSIS\'!C{i}="S1",\'08_MERGED_ANALYSIS\'!J{i}<>"",'
     '\'08_MERGED_ANALYSIS\'!J{i}>=0,\'08_MERGED_ANALYSIS\'!J{i}<=3600),\'08_MERGED_ANALYSIS\'!J{i},"")'),
    ("completion_s2_clean",
     '=IF(AND(\'08_MERGED_ANALYSIS\'!C{i}="S2",\'08_MERGED_ANALYSIS\'!J{i}<>"",'
     '\'08_MERGED_ANALYSIS\'!J{i}>=0,\'08_MERGED_ANALYSIS\'!J{i}<=3600),\'08_MERGED_ANALYSIS\'!J{i},"")'),
    ("latency_s1", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S1",\'08_MERGED_ANALYSIS\'!K{i},"")'),
    ("latency_s2", '=IF(\'08_MERGED_ANALYSIS\'!C{i}="S2",\'08_MERGED_ANALYSIS\'!K{i},"")'),
    ("asq_mean_success1", '=IF(\'08_MERGED_ANALYSIS\'!H{i}=1,\'08_MERGED_ANALYSIS\'!G{i},"")'),
    ("asq_mean_success0", '=IF(\'08_MERGED_ANALYSIS\'!H{i}=0,\'08_MERGED_ANALYSIS\'!G{i},"")'),
    ("asq3_for_citation", '=IF(\'08_MERGED_ANALYSIS\'!L{i}<>"",\'08_MERGED_ANALYSIS\'!F{i},"")'),
    ("citation_as_number",
     '=IF(\'08_MERGED_ANALYSIS\'!L{i}<>"",IF(\'08_MERGED_ANALYSIS\'!L{i},1,0),"")'),
]
HELPER_START_COL = 14  # column N


def build_09_statistical_summary(wb, df: pd.DataFrame) -> dict:
    """Returns the Python-computed values (for cross-check / report use).

    Uses plain (non-array) formulas throughout: every scenario/group-conditional
    aggregate is computed via a per-row helper column (see HELPER_COLS) rather
    than a legacy CSE array formula, since there is no way to verify array-formula
    recalculation behavior in this environment and a silently-broken formula in a
    research-integrity deliverable is worse than a slightly wider sheet.
    """
    ws = wb.create_sheet("09_STATISTICAL_SUMMARY")
    ws.column_dimensions["A"].width = 4
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 18

    merged_range = "'08_MERGED_ANALYSIS'"
    n = len(df)

    # -- Write helper columns (N onward), one row per observation, mirroring
    # 08_MERGED_ANALYSIS's row order exactly.
    helper_col_letter = {}
    for offset, (name, _tmpl) in enumerate(HELPER_COLS):
        col_idx = HELPER_START_COL + offset
        letter = get_column_letter(col_idx)
        helper_col_letter[name] = letter
        ws.cell(row=1, column=col_idx, value=f"helper: {name}")
    for i in range(2, n + 2):
        for offset, (name, tmpl) in enumerate(HELPER_COLS):
            col_idx = HELPER_START_COL + offset
            ws.cell(row=i, column=col_idx, value=tmpl.format(i=i))
    ws.cell(row=1, column=HELPER_START_COL).font = Font(italic=True, size=8)
    for offset in range(len(HELPER_COLS)):
        ws.cell(row=1, column=HELPER_START_COL + offset).font = Font(italic=True, size=8)

    def hr(name: str) -> str:
        """Helper column range, e.g. 'N2:N44'."""
        letter = helper_col_letter[name]
        return f"{letter}2:{letter}{n + 1}"

    asq_col = f"{merged_range}!$G$2:$G${n + 1}"
    s1_mask = f"{merged_range}!$C$2:$C${n + 1}"
    task_col = f"{merged_range}!$H$2:$H${n + 1}"
    lat_col = f"{merged_range}!$K$2:$K${n + 1}"

    r = 1
    ws.cell(row=r, column=1, value="A. Descriptive Statistics -- ASQ mean score").font = TITLE_FONT
    r += 2
    headers = ["Statistic", "Overall", "S1", "S2"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h)
    style_header_row(ws, r, 4)
    hdr_row = r
    r += 1
    stat_rows = {}
    stats_defs = [
        ("n", f"=COUNT({asq_col})", f'=COUNTIFS({s1_mask},"S1",{asq_col},"<>")',
         f'=COUNTIFS({s1_mask},"S2",{asq_col},"<>")'),
        ("Mean", f"=AVERAGE({asq_col})",
         f'=AVERAGEIFS({asq_col},{s1_mask},"S1")', f'=AVERAGEIFS({asq_col},{s1_mask},"S2")'),
        ("Median", f"=MEDIAN({asq_col})", None, None),
        ("SD", f"=STDEV.S({asq_col})", None, None),
        ("Min", f"=MIN({asq_col})", None, None),
        ("Max", f"=MAX({asq_col})", None, None),
    ]
    for label, overall_f, s1_f, s2_f in stats_defs:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=overall_f)
        if s1_f:
            ws.cell(row=r, column=3, value=s1_f)
        if s2_f:
            ws.cell(row=r, column=4, value=s2_f)
        stat_rows[label] = r
        r += 1
    n_row, mean_row, sd_row = stat_rows["n"], stat_rows["Mean"], stat_rows["SD"]
    ws.cell(row=r, column=1, value="95% CI lower")
    ws.cell(row=r, column=2,
            value=f"=B{mean_row}-CONFIDENCE.T(0.05,B{sd_row},B{n_row})")
    ci_lo_row = r
    r += 1
    ws.cell(row=r, column=1, value="95% CI upper")
    ws.cell(row=r, column=2,
            value=f"=B{mean_row}+CONFIDENCE.T(0.05,B{sd_row},B{n_row})")
    ci_hi_row = r
    r += 2

    # B. Cronbach's alpha per scenario -- live formula, via helper columns (no
    # array-entry required: VAR.S over a plain range naturally ignores blanks).
    ws.cell(row=r, column=1, value="B. Cronbach's Alpha per Scenario (3 ASQ items, indicative)").font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=1, value="Scenario")
    ws.cell(row=r, column=2, value="n")
    ws.cell(row=r, column=3, value="Cronbach's alpha")
    style_header_row(ws, r, 3)
    r += 1
    for scen, q_names, total_name in [
        ("S1", ("q1_s1", "q2_s1", "q3_s1"), "total_s1"),
        ("S2", ("q1_s2", "q2_s2", "q3_s2"), "total_s2"),
    ]:
        ws.cell(row=r, column=1, value=scen)
        ws.cell(row=r, column=2, value=f"=COUNT({hr(q_names[0])})")
        var_sum = "+".join(f"VAR.S({hr(name)})" for name in q_names)
        alpha_formula = f"=(3/2)*(1-({var_sum})/VAR.S({hr(total_name)}))"
        ws.cell(row=r, column=3, value=alpha_formula)
        r += 1
    r += 1
    ws.cell(row=r - 1, column=1,
            value="Formula: Cronbach's alpha = (k/(k-1))*(1 - sum(item variances)/variance(total "
                  "score)), computed via the helper columns in N:AC (one row per observation, "
                  "scenario-filtered) -- plain VAR.S formulas, no array/CSE entry required.").alignment = WRAP
    r += 1

    # C. Task success rate -- live formula
    ws.cell(row=r, column=1, value="C. Task Success Rate per Scenario "
                                    "(success = answer_status in {answered, verified})").font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=1, value="Scenario")
    ws.cell(row=r, column=2, value="n with linked turn")
    ws.cell(row=r, column=3, value="n success")
    ws.cell(row=r, column=4, value="success rate")
    style_header_row(ws, r, 4)
    r += 1
    for scen in ["S1", "S2"]:
        ws.cell(row=r, column=1, value=scen)
        n_linked = f'=COUNTIFS({s1_mask},"{scen}",{task_col},"<>")'
        n_succ = f'=COUNTIFS({s1_mask},"{scen}",{task_col},1)'
        ws.cell(row=r, column=2, value=n_linked)
        ws.cell(row=r, column=3, value=n_succ)
        ws.cell(row=r, column=4, value=f"=C{r}/B{r}")
        r += 1
    r += 1

    # D. Completion time & latency -- live formula, via helper columns
    ws.cell(row=r, column=1, value="D. Completion Time (sec, proxy) and Latency (ms) per Scenario").font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=1, value="Scenario")
    ws.cell(row=r, column=2, value="Median completion_time_sec (clean)")
    ws.cell(row=r, column=3, value="Mean latency_ms")
    ws.cell(row=r, column=4, value="Median latency_ms")
    style_header_row(ws, r, 4)
    r += 1
    for scen, completion_name, latency_name in [
        ("S1", "completion_s1_clean", "latency_s1"), ("S2", "completion_s2_clean", "latency_s2"),
    ]:
        ws.cell(row=r, column=1, value=scen)
        ws.cell(row=r, column=2, value=f"=MEDIAN({hr(completion_name)})")
        ws.cell(row=r, column=3, value=f"=AVERAGE({hr(latency_name)})")
        ws.cell(row=r, column=4, value=f"=MEDIAN({hr(latency_name)})")
        r += 1
    r += 1

    # E. Subjective vs objective (ASQ mean by task_success group) -- live formula
    ws.cell(row=r, column=1, value="E. ASQ Mean by Task Success Group (subjective vs. objective)").font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=1, value="Group")
    ws.cell(row=r, column=2, value="n")
    ws.cell(row=r, column=3, value="Mean ASQ")
    ws.cell(row=r, column=4, value="SD ASQ")
    style_header_row(ws, r, 4)
    r += 1
    for grp, helper_name in [("task_success = 1", "asq_mean_success1"),
                              ("task_success = 0", "asq_mean_success0")]:
        ws.cell(row=r, column=1, value=grp)
        ws.cell(row=r, column=2, value=f"=COUNT({hr(helper_name)})")
        ws.cell(row=r, column=3, value=f"=AVERAGE({hr(helper_name)})")
        ws.cell(row=r, column=4, value=f"=STDEV.S({hr(helper_name)})")
        r += 1
    r += 1

    # F. Correlations
    ws.cell(row=r, column=1, value="F. Correlations").font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=1, value="Pair")
    ws.cell(row=r, column=2, value="n")
    ws.cell(row=r, column=3, value="r / rho")
    ws.cell(row=r, column=4, value="p-value")
    ws.cell(row=r, column=5, value="method")
    style_header_row(ws, r, 5)
    r += 1
    ws.cell(row=r, column=1, value="ASQ mean vs latency_ms (Pearson, live formula)")
    ws.cell(row=r, column=2, value=f'=COUNT({lat_col})')
    ws.cell(row=r, column=3, value=f"=CORREL({asq_col},{lat_col})")
    ws.cell(row=r, column=4,
            value=f"=T.DIST.2T(ABS(C{r})*SQRT(B{r}-2)/SQRT(1-C{r}^2),B{r}-2)")
    ws.cell(row=r, column=5, value="Pearson (live formula)")
    r += 1
    ws.cell(row=r, column=1, value="ASQ mean vs task_success (point-biserial, live formula)")
    ws.cell(row=r, column=2, value=f'=COUNT({task_col})')
    ws.cell(row=r, column=3, value=f"=CORREL({asq_col},{task_col})")
    ws.cell(row=r, column=4,
            value=f"=T.DIST.2T(ABS(C{r})*SQRT(B{r}-2)/SQRT(1-C{r}^2),B{r}-2)")
    ws.cell(row=r, column=5, value="point-biserial = Pearson on 0/1 (live formula)")
    r += 1
    ws.cell(row=r, column=1, value="ASQ_3 vs has_correct_citation (point-biserial, live formula)")
    ws.cell(row=r, column=2, value=f"=COUNT({hr('citation_as_number')})")
    ws.cell(row=r, column=3, value=f"=CORREL({hr('asq3_for_citation')},{hr('citation_as_number')})")
    ws.cell(row=r, column=4,
            value=f"=T.DIST.2T(ABS(C{r})*SQRT(B{r}-2)/SQRT(1-C{r}^2),B{r}-2)")
    ws.cell(row=r, column=5, value="point-biserial = Pearson on 0/1, via helper cols (live formula)")
    r += 1
    ws.cell(row=r, column=1,
            value="Spearman rho: ASQ mean vs latency_ms (Python-computed, see note)")
    spearman_row = r
    r += 1
    r += 1
    ws.cell(row=r, column=1,
            value="Note: Spearman correlation is reported as a Python-computed value (scipy, "
                  "same method as build_asq_scenario_excel_2026-08-02.py) rather than a live "
                  "formula, because ASQ is an ordinal Likert measure and rank-based Spearman is "
                  "more appropriate than Pearson for it; Pearson/point-biserial above are kept "
                  "live since they match scipy exactly.").alignment = WRAP
    r += 2

    # G. Wilcoxon
    ws.cell(row=r, column=1, value="G. Wilcoxon Signed-Rank Test, ASQ mean S1 vs S2 "
                                    "(Python-computed, see note)").font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=1, value="paired_n")
    ws.cell(row=r, column=2, value="median_s1")
    ws.cell(row=r, column=3, value="median_s2")
    ws.cell(row=r, column=4, value="W statistic")
    ws.cell(row=r, column=5, value="p-value")
    style_header_row(ws, r, 5)
    r += 1
    wilcoxon_row = r
    r += 2
    ws.cell(row=r, column=1,
            value="Note: computed once in Python (scipy.stats.wilcoxon, exact method) because "
                  "Excel has no native exact Wilcoxon signed-rank function; a hand-built "
                  "normal-approximation formula would risk disagreeing with the standard method "
                  "for reasons unrelated to the data (this dataset has 13/21 tied S1=S2 pairs, "
                  "exactly where hand-rolled and exact implementations diverge most). See "
                  "build_research_dataset_excel_2026-08-10.py::compute_wilcoxon.").alignment = WRAP
    r += 2

    # H. Perception-performance gap (continuous framing, no fabricated threshold)
    ws.cell(row=r, column=1, value="H. Perception-Performance Gap "
                                    "(continuous framing; no threshold imposed)").font = TITLE_FONT
    r += 2
    ws.cell(row=r, column=1,
            value="Primary comparison: see section E above (mean ASQ by task_success group). No "
                  "high/low ASQ threshold is imposed here because none is specified or validated "
                  "in the source methodology -- inventing one would be an integrity risk in this "
                  "dataset. If a categorical exploratory view is wanted later, add it as a clearly "
                  "labeled secondary table with the threshold stated explicitly next to it.").alignment = WRAP
    r += 2

    # I. Reconciliation table
    ws.cell(row=r, column=1, value="I. Reconciliation vs. Values Reported as Already Published").font = TITLE_FONT
    r += 2
    recon_headers = ["statistic", "reported_value", "recomputed_value", "difference", "status",
                      "source_of_reported_value"]
    for j, h in enumerate(recon_headers, start=1):
        ws.cell(row=r, column=j, value=h)
    style_header_row(ws, r, len(recon_headers))
    recon_hdr_row = r
    r += 1
    ws.cell(row=r, column=1, value="ASQ overall mean")
    ws.cell(row=r, column=2, value=5.76)
    ws.cell(row=r, column=3, value=f"=ROUND(B{mean_row},2)")
    ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
    ws.cell(row=r, column=5, value=f'=IF(ABS(D{r})<0.01,"MATCH",IF(ABS(D{r})<0.1,"ROUNDING_DIFFERENCE","MISMATCH"))')
    ws.cell(row=r, column=6, value="reported by user as already-published in the paper")
    r += 1
    ws.cell(row=r, column=1, value="n observations")
    ws.cell(row=r, column=2, value=43)
    ws.cell(row=r, column=3, value=f"=B{n_row}")
    ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
    ws.cell(row=r, column=5, value=f'=IF(D{r}=0,"MATCH","MISMATCH")')
    ws.cell(row=r, column=6, value="reported by user as already-published in the paper")
    r += 1
    ws.cell(row=r, column=1, value="n participants")
    ws.cell(row=r, column=2, value=22)
    ws.cell(row=r, column=3, value="=COUNTA('02_PARTICIPANTS'!A2:A100)")
    ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
    ws.cell(row=r, column=5, value=f'=IF(D{r}=0,"MATCH","MISMATCH")')
    ws.cell(row=r, column=6, value="reported by user as already-published in the paper")
    r += 1
    other_stats = [
        "Cronbach's alpha S1", "Cronbach's alpha S2", "Task success rate S1",
        "Task success rate S2", "Completion time median S1 (sec)",
        "Completion time median S2 (sec)", "Latency mean S1 (ms)", "Latency mean S2 (ms)",
        "Spearman rho ASQ vs latency", "Wilcoxon p (ASQ S1 vs S2)",
    ]
    for stat_name in other_stats:
        ws.cell(row=r, column=1, value=stat_name)
        ws.cell(row=r, column=2, value="SOURCE_NOT_FOUND")
        ws.cell(row=r, column=3, value="see corresponding section above")
        ws.cell(row=r, column=4, value="")
        ws.cell(row=r, column=5, value="SOURCE_NOT_FOUND")
        ws.cell(row=r, column=6,
                value="no independently-stated paper value was supplied for this statistic; not "
                      "back-filled with a fabricated target")
        r += 1
    recon_last_row = r - 1

    autosize(ws, max_width=60)

    # Compute Python-side values for embedding + cross-check
    joined = df.rename(columns={"asq_1": "asq_1", "asq_2": "asq_2", "asq_3": "asq_3"})
    desc_overall = descriptive(joined["average_score"])
    cronbach_s1 = cronbach_alpha(joined[joined["scenario"] == "S1"][["asq_1", "asq_2", "asq_3"]])
    cronbach_s2 = cronbach_alpha(joined[joined["scenario"] == "S2"][["asq_1", "asq_2", "asq_3"]])
    task_success_df = compute_task_success(joined)
    completion_df = compute_completion_time(joined)
    latency_df = compute_latency(joined)
    correlations = compute_correlations(joined)
    wilcoxon = compute_wilcoxon(joined)

    # Write Spearman value
    ws.cell(row=spearman_row, column=2, value=nv(correlations["asq_vs_latency"]["n"]))
    ws.cell(row=spearman_row, column=3, value=nv(correlations["asq_vs_latency"]["spearman_r"]))
    ws.cell(row=spearman_row, column=4, value=nv(correlations["asq_vs_latency"]["spearman_p"]))
    ws.cell(row=spearman_row, column=5, value="Spearman (Python/scipy, computed once)")

    # Write Wilcoxon values
    ws.cell(row=wilcoxon_row, column=1, value=wilcoxon["n_pairs"])
    ws.cell(row=wilcoxon_row, column=2,
            value=nv(joined[joined["scenario"] == "S1"]["average_score"].median()))
    ws.cell(row=wilcoxon_row, column=3,
            value=nv(joined[joined["scenario"] == "S2"]["average_score"].median()))
    ws.cell(row=wilcoxon_row, column=4, value=nv(wilcoxon["statistic"]))
    ws.cell(row=wilcoxon_row, column=5, value=nv(wilcoxon["p"]))

    return {
        "desc_overall": desc_overall, "cronbach_s1": cronbach_s1, "cronbach_s2": cronbach_s2,
        "task_success_df": task_success_df, "completion_df": completion_df,
        "latency_df": latency_df, "correlations": correlations, "wilcoxon": wilcoxon,
    }


def build_10_data_quality(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("10_DATA_QUALITY")
    headers = ["check_name", "status", "affected_rows", "description"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    n = len(df)
    obs_range = f"'03_OBSERVATIONS'!$A$2:$A${n + 1}"
    part_range = f"'03_OBSERVATIONS'!$B$2:$B${n + 1}"
    scen_range = f"'03_OBSERVATIONS'!$C$2:$C${n + 1}"
    asq1_range = f"'04_ASQ_RESPONSES'!$D$2:$D${n + 1}"
    asq3_range = f"'04_ASQ_RESPONSES'!$F$2:$F${n + 1}"
    lat_range = f"'06_SYSTEM_METRICS'!$D$2:$D${n + 1}"
    ctime_range = f"'05_TASK_PERFORMANCE'!$F$2:$F${n + 1}"
    task_range = f"'05_TASK_PERFORMANCE'!$D$2:$D${n + 1}"

    checks = [
        ("duplicate observation_id", f"=SUMPRODUCT((COUNTIF({obs_range},{obs_range})>1)*1)",
         "Count of observation_id values that appear more than once (expect 0)."),
        ("duplicate participant-scenario pair",
         f"=SUMPRODUCT((COUNTIFS({part_range},{part_range},{scen_range},{scen_range})>1)*1)",
         "Count of (participant_id, scenario_id) combinations that appear more than once (expect 0)."),
        ("participant without any observation",
         f"=SUMPRODUCT((COUNTIF({part_range},'02_PARTICIPANTS'!$A$2:$A$23)=0)*1)",
         "Participants in 02_PARTICIPANTS with zero rows in 03_OBSERVATIONS (expect 0)."),
        ("ASQ item out of 1-7 range",
         f"=SUMPRODUCT(((({asq1_range}<1)+({asq1_range}>7))>0)*1)+SUMPRODUCT(((({asq3_range}<1)+({asq3_range}>7))>0)*1)",
         "Any ASQ item value outside the valid 1-7 Likert range (expect 0)."),
        ("negative latency_ms",
         f"=SUMPRODUCT((({lat_range}<>\"\")*({lat_range}<0))*1)",
         "latency_ms values below zero (expect 0)."),
        ("negative completion_time_sec",
         f"=SUMPRODUCT((({ctime_range}<>\"\")*({ctime_range}<0))*1)",
         "completion_time_sec values below zero -- a genuine data quirk (1 row expected: "
         "participant with a negative proxy duration, source-flagged completion_time_valid=False)."),
        (">1h outlier completion_time_sec",
         f"=SUMPRODUCT((({ctime_range}<>\"\")*({ctime_range}>3600))*1)",
         "completion_time_sec values over 1 hour -- genuine outliers (2 rows expected, "
         "source-flagged completion_time_outlier=True)."),
        ("invalid scenario_id",
         f'=SUMPRODUCT((({scen_range}<>"S1")*({scen_range}<>"S2"))*1)',
         "scenario_id values outside {S1, S2} (expect 0)."),
        ("missing ASQ (incomplete item set)",
         f"=COUNTIF('04_ASQ_RESPONSES'!$G$2:$G${n + 1},\"\")",
         "Rows where asq_mean is blank, i.e. at least one of the 3 ASQ items is missing "
         "(expect 0; all 43 rows have full ASQ item sets)."),
        ("missing task_success",
         f'=COUNTIF({task_range},"")',
         "Rows with blank task_success -- expected 1 (row with no linked production trace_id)."),
        ("missing latency",
         f'=COUNTIF({lat_range},"")',
         "Rows with blank latency_ms -- expected 1 (same no-linked-turn row)."),
        ("incomplete scenario pair (participant missing S1 or S2)",
         f"=SUMPRODUCT((COUNTIF({part_range},'02_PARTICIPANTS'!$A$2:$A$23)<2)*1)",
         "Participants with fewer than 2 observation rows -- expected 1 (participant missing S2)."),
    ]

    expected_warn = {
        "negative completion_time_sec", ">1h outlier completion_time_sec",
        "missing task_success", "missing latency",
        "incomplete scenario pair (participant missing S1 or S2)",
    }

    r = 2
    for name, formula, desc in checks:
        ws.cell(row=r, column=1, value=name)
        affected_col = 3
        ws.cell(row=r, column=affected_col, value=formula)
        status_formula = (f'=IF(C{r}=0,"PASS",IF(C{r}>0,'
                           f'"{"WARNING" if name in expected_warn else "FAIL"}","PASS"))')
        ws.cell(row=r, column=2, value=status_formula)
        ws.cell(row=r, column=4, value=desc)
        r += 1
    last_row = r - 1
    add_table(ws, "tbl_10_data_quality", f"A1:D{last_row}")
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    autosize(ws, max_width=90)
    ws.cell(row=1, column=1).alignment = WRAP
    ws.column_dimensions["D"].width = 90


# ---------------------------------------------------------------------------
# Codebook completeness assertion
# ---------------------------------------------------------------------------

def assert_codebook_matches(written: dict[str, list[str]]) -> None:
    """Diff the (sheet, column) pairs actually written vs. what CODEBOOK documents."""
    documented = {(c.sheet, c.variable_name) for c in CODEBOOK if c.sheet in written}
    actual = set()
    for sheet, cols in written.items():
        if sheet not in {"09_STATISTICAL_SUMMARY", "10_DATA_QUALITY"}:
            for col in cols:
                actual.add((sheet, col))
    missing_from_codebook = actual - documented
    if missing_from_codebook:
        raise AssertionError(f"Columns written but not documented in codebook: {sorted(missing_from_codebook)}")
    extra_in_codebook = documented - actual
    if extra_in_codebook:
        raise AssertionError(f"Codebook documents columns never written: {sorted(extra_in_codebook)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-csv", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    args = parser.parse_args()

    df = load_and_shape(args.input_csv)
    assert len(df) == 43, f"expected 43 observations, got {len(df)}"
    assert df["participant_id"].nunique() == 22, f"expected 22 participants, got {df['participant_id'].nunique()}"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_00_readme(wb, df)
    build_01_codebook(wb)
    build_02_participants(wb, df)
    build_03_observations(wb, df)
    build_04_asq(wb, df)
    build_05_task_performance(wb, df)
    build_06_system_metrics(wb, df)
    build_07_citation(wb, df)
    build_08_merged(wb, df)
    python_stats = build_09_statistical_summary(wb, df)
    build_10_data_quality(wb, df)

    written_columns = {
        "02_PARTICIPANTS": ["participant_id", "participant_code", "participant_group",
                             "completed_s1", "completed_s2", "total_observations", "notes"],
        "03_OBSERVATIONS": ["observation_id", "participant_id", "scenario_id", "scenario_order",
                             "session_id", "timestamp_start", "timestamp_end", "query_text",
                             "response_id", "task_status", "data_source", "missing_flag",
                             "missing_reason"],
        "04_ASQ_RESPONSES": ["observation_id", "participant_id", "scenario_id", "asq_q1_ease",
                              "asq_q2_time", "asq_q3_information_support", "asq_mean",
                              "asq_complete", "asq_notes"],
        "05_TASK_PERFORMANCE": ["observation_id", "participant_id", "scenario_id", "task_success",
                                 "task_status", "completion_time_sec", "completion_time_source",
                                 "success_evidence", "evaluator_notes"],
        "06_SYSTEM_METRICS": ["observation_id", "participant_id", "scenario_id", "latency_ms",
                               "retrieval_latency_ms", "llm_latency_ms", "retrieved_chunks",
                               "graph_nodes_retrieved", "graph_relationships", "token_input",
                               "token_output", "model_name", "request_status", "missing_reason"],
        "07_CITATION_EVALUATION": ["observation_id", "participant_id", "scenario_id",
                                    "n_citations", "n_valid_citations", "has_correct_citation",
                                    "citation_evaluation_method", "evaluator_notes"],
        "08_MERGED_ANALYSIS": ["observation_id", "participant_id", "scenario_id", "asq_q1_ease",
                                "asq_q2_time", "asq_q3_information_support", "asq_mean",
                                "task_success", "task_status", "completion_time_sec", "latency_ms",
                                "has_correct_citation", "complete_case_asq",
                                "complete_case_performance", "paired_scenario_available"],
    }
    assert_codebook_matches(written_columns)

    order = ["00_README", "01_CODEBOOK", "02_PARTICIPANTS", "03_OBSERVATIONS",
             "04_ASQ_RESPONSES", "05_TASK_PERFORMANCE", "06_SYSTEM_METRICS",
             "07_CITATION_EVALUATION", "08_MERGED_ANALYSIS", "09_STATISTICAL_SUMMARY",
             "10_DATA_QUALITY"]
    wb._sheets = [wb[name] for name in order]

    args.out_dir.mkdir(parents=True, exist_ok=True)
    out_path = args.out_dir / "research_dataset_perception_performance.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")

    # Print Python-computed cross-check values for the reconstruction report.
    print("\n--- Python-computed cross-check values ---")
    print(f"n_participants={df['participant_id'].nunique()} n_observations={len(df)}")
    print(f"ASQ overall: {python_stats['desc_overall']}")
    print(f"Cronbach S1={python_stats['cronbach_s1']:.4f} S2={python_stats['cronbach_s2']:.4f}")
    print(python_stats["task_success_df"])
    print(python_stats["completion_df"])
    print(python_stats["latency_df"])
    print(python_stats["correlations"])
    print(python_stats["wilcoxon"])


if __name__ == "__main__":
    main()
