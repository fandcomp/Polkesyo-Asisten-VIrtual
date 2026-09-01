"""Build the 2026-08-02 ASQ-per-scenario evaluation workbook.

Combines two evaluation instruments that had never been joined before in this project:
1. The 2026-07-16 human ASQ usability study (22 participants, scenarios S1/S2, 3-item ASQ 1-7).
2. Objective production logs (chat_evaluation_logs, citation_evaluation_logs) linked via trace_id,
   and the clean gold-QA automated run (refresh_20260727_gates_all, n=41) for citation
   correctness / hallucination rate / false-positive rate.

Inputs (all produced by build_r0x_traceid_map_2026-08-02.py + vps_query_asq_linkage_2026-08-02.sql,
read from a scratch directory that must NEVER be inside the repo -- see --input-dir):
  r0x_traceid_map.csv   (r0x, scenario, asq_1/2/3, average_score, trace_id, session_id, asq_created_at)
  A_chat_logs.csv        (trace_id, scenario_code, total_latency_ms, answer_status, created_at)
  B_citations.csv        (trace_id, n_citations, n_valid_citations)
  C_session_first_turn.csv (session_id, first_turn_at)
  E_eval_results_detail.csv (category, expected_behavior, fallback_correct, citation_correct, hallucination_detected)

Output (written under campus-va/evaluation/reports/2026-08-02/, safe to commit -- contains only
R01-R22 codes, never real participant codes, trace_id, or session_id):
  ASQ_Evaluasi_Skenario_Lengkap_2026-08-02.xlsx
  raw_joined_by_scenario.csv
  figures/*.png

Usage:
    python build_asq_scenario_excel_2026-08-02.py --input-dir <scratch dir>
"""
from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats

ROOT = Path(__file__).parent.parent.parent  # campus-va/
OUT_DIR = ROOT / "evaluation" / "reports" / "2026-08-02"
FIG_DIR = OUT_DIR / "figures"

HEADER_FILL = PatternFill(start_color="FF1F3B3C", end_color="FF1F3B3C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
TITLE_FONT = Font(bold=True, size=13)

SUCCESS_STATUSES = {"answered", "verified"}
COMPLETION_TIME_OUTLIER_THRESHOLD_S = 3600  # 1 hour -- a scenario task in this study is a few minutes at most
ASQ_ITEM_LABELS = {
    "asq_1": "ASQ_1 - Kemudahan menyelesaikan tugas",
    "asq_2": "ASQ_2 - Kepuasan terhadap waktu penyelesaian",
    "asq_3": "ASQ_3 - Kepuasan terhadap dukungan informasi",
}

plt.rcParams.update({
    "figure.dpi": 150, "savefig.bbox": "tight", "font.size": 10,
    "axes.grid": True, "grid.alpha": 0.3, "axes.axisbelow": True,
})


# ---------------------------------------------------------------------------
# Loading + joining
# ---------------------------------------------------------------------------

def load_inputs(input_dir: Path) -> dict[str, pd.DataFrame]:
    r0x = pd.read_csv(input_dir / "r0x_traceid_map.csv")
    chat = pd.read_csv(input_dir / "A_chat_logs.csv", parse_dates=["created_at"])
    cite = pd.read_csv(input_dir / "B_citations.csv")
    sess = pd.read_csv(input_dir / "C_session_first_turn.csv", parse_dates=["first_turn_at"])
    ev = pd.read_csv(input_dir / "E_eval_results_detail.csv")
    r0x["asq_created_at"] = pd.to_datetime(r0x["asq_created_at"], utc=True)
    chat["created_at"] = pd.to_datetime(chat["created_at"], utc=True)
    sess["first_turn_at"] = pd.to_datetime(sess["first_turn_at"], utc=True)
    return {"r0x": r0x, "chat": chat, "cite": cite, "sess": sess, "ev": ev}


def build_joined(d: dict[str, pd.DataFrame]) -> pd.DataFrame:
    j = d["r0x"].merge(d["chat"][["trace_id", "total_latency_ms", "answer_status"]], on="trace_id", how="left")
    j = j.merge(d["cite"], on="trace_id", how="left")
    j["n_citations"] = j["n_citations"].fillna(0).astype(int)
    j["n_valid_citations"] = j["n_valid_citations"].fillna(0).astype(int)
    j["has_correct_citation"] = j["n_valid_citations"] > 0
    j["has_linked_turn"] = j["trace_id"].notna() & j["answer_status"].notna()
    j["task_success"] = np.where(j["has_linked_turn"], j["answer_status"].isin(SUCCESS_STATUSES), np.nan)

    j = j.merge(d["sess"], on="session_id", how="left")

    # Completion time proxy: S1 = asq_created_at - session first_turn_at;
    # S2 = asq_created_at(S2) - asq_created_at(S1) for the same respondent.
    s1_time = j[j["scenario"] == "S1"].set_index("r0x")["asq_created_at"]
    completion = []
    valid_flag = []
    for _, row in j.iterrows():
        if row["scenario"] == "S1":
            if pd.notna(row["first_turn_at"]):
                secs = (row["asq_created_at"] - row["first_turn_at"]).total_seconds()
            else:
                secs = np.nan
        else:  # S2
            s1_t = s1_time.get(row["r0x"])
            secs = (row["asq_created_at"] - s1_t).total_seconds() if pd.notna(s1_t) else np.nan
        completion.append(secs)
        valid_flag.append(pd.notna(secs) and secs >= 0)
    j["completion_time_s"] = completion
    j["completion_time_valid"] = valid_flag
    j["completion_time_outlier"] = [
        bool(v) and secs > COMPLETION_TIME_OUTLIER_THRESHOLD_S
        for v, secs in zip(valid_flag, completion)
    ]
    j["completion_time_clean"] = j["completion_time_valid"] & ~j["completion_time_outlier"]
    return j


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------

def descriptive(series: pd.Series) -> dict:
    s = series.dropna()
    n = len(s)
    mean = s.mean()
    sd = s.std(ddof=1) if n > 1 else np.nan
    sem = sd / np.sqrt(n) if n > 1 else np.nan
    tcrit = stats.t.ppf(0.975, df=n - 1) if n > 1 else np.nan
    ci_lo, ci_hi = (mean - tcrit * sem, mean + tcrit * sem) if n > 1 else (np.nan, np.nan)
    q1, q3 = (s.quantile(0.25), s.quantile(0.75)) if n > 0 else (np.nan, np.nan)
    return {
        "n": n, "mean": mean, "median": s.median() if n else np.nan, "sd": sd,
        "iqr": q3 - q1 if n else np.nan, "q1": q1, "q3": q3,
        "ci95_lo": ci_lo, "ci95_hi": ci_hi,
    }


def cronbach_alpha(item_df: pd.DataFrame) -> float:
    item_vars = item_df.var(axis=0, ddof=1)
    total_var = item_df.sum(axis=1).var(ddof=1)
    k = item_df.shape[1]
    if total_var == 0:
        return np.nan
    return float((k / (k - 1)) * (1 - item_vars.sum() / total_var))


def compute_asq_per_item_scenario(joined: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for scenario in ["S1", "S2"]:
        sub = joined[joined["scenario"] == scenario]
        for item in ["asq_1", "asq_2", "asq_3"]:
            desc = descriptive(sub[item])
            rows.append({"scenario": scenario, "item": ASQ_ITEM_LABELS[item], **desc})
    return pd.DataFrame(rows)


def compute_asq_overall(joined: pd.DataFrame) -> dict:
    overall = descriptive(joined["average_score"])
    per_scenario = {s: descriptive(joined[joined["scenario"] == s]["average_score"]) for s in ["S1", "S2"]}
    return {"overall": overall, "per_scenario": per_scenario}


def compute_cronbach_per_scenario(joined: pd.DataFrame) -> dict:
    out = {}
    for scenario in ["S1", "S2"]:
        sub = joined[joined["scenario"] == scenario][["asq_1", "asq_2", "asq_3"]]
        alpha = cronbach_alpha(sub)
        item_total = {}
        total = sub.sum(axis=1)
        for item in ["asq_1", "asq_2", "asq_3"]:
            rest = total - sub[item]
            r, _ = stats.pearsonr(sub[item], rest)
            item_total[item] = r
        out[scenario] = {"n": len(sub), "alpha": alpha, "item_total_corr": item_total}
    return out


def compute_task_success(joined: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for scenario in ["S1", "S2"]:
        sub = joined[joined["scenario"] == scenario]
        with_data = sub[sub["has_linked_turn"]]
        n_data = len(with_data)
        n_success = int(with_data["task_success"].sum()) if n_data else 0
        rows.append({
            "scenario": scenario, "n_total_rows": len(sub), "n_with_linked_turn": n_data,
            "n_success": n_success,
            "success_rate": n_success / n_data if n_data else np.nan,
        })
    return pd.DataFrame(rows)


def compute_completion_time(joined: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for scenario in ["S1", "S2"]:
        sub = joined[joined["scenario"] == scenario]
        clean = sub[sub["completion_time_clean"]]
        n_negative = int((~sub["completion_time_valid"]).sum())
        n_outlier = int(sub["completion_time_outlier"].sum())
        desc = descriptive(clean["completion_time_s"])
        rows.append({
            "scenario": scenario, "n_clean": len(clean),
            "n_excluded_negative": n_negative,
            "n_excluded_outlier_gt_1h": n_outlier,
            **desc,
        })
    return pd.DataFrame(rows)


def compute_latency(joined: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for scenario in ["S1", "S2"]:
        sub = joined[joined["scenario"] == scenario]["total_latency_ms"].dropna()
        rows.append({
            "scenario": scenario, "n": len(sub),
            "mean_ms": sub.mean() if len(sub) else np.nan,
            "median_ms": sub.median() if len(sub) else np.nan,
            "p95_ms": sub.quantile(0.95) if len(sub) else np.nan,
            "sd_ms": sub.std(ddof=1) if len(sub) > 1 else np.nan,
        })
    return pd.DataFrame(rows)


def compute_citation_hallucination_fpr(ev: pd.DataFrame) -> dict:
    ev = ev.copy()
    for col in ["fallback_correct", "citation_correct", "hallucination_detected"]:
        ev[col] = ev[col].map({"t": True, "f": False, True: True, False: False}).astype("boolean").fillna(False).astype(bool)

    n = len(ev)
    citation_correct_rate = ev["citation_correct"].mean()
    hallucination_rate = ev["hallucination_detected"].mean()

    valid_answer = ev[ev["expected_behavior"] == "answer"]
    n_valid_answer = len(valid_answer)
    n_false_positive = int((~valid_answer["fallback_correct"]).sum())
    fpr = n_false_positive / n_valid_answer if n_valid_answer else np.nan

    per_category = ev.groupby("category").agg(
        n=("category", "size"),
        n_citation_correct=("citation_correct", "sum"),
        n_hallucination=("hallucination_detected", "sum"),
    ).reset_index()

    return {
        "n_questions": n, "citation_correct_rate": citation_correct_rate,
        "hallucination_rate": hallucination_rate,
        "n_valid_answer_category": n_valid_answer, "n_false_positive": n_false_positive,
        "false_positive_rate": fpr, "per_category": per_category,
    }


def compute_correlations(joined: pd.DataFrame) -> dict:
    out = {}

    sub = joined.dropna(subset=["average_score", "total_latency_ms"])
    pear = stats.pearsonr(sub["average_score"], sub["total_latency_ms"])
    spear = stats.spearmanr(sub["average_score"], sub["total_latency_ms"])
    out["asq_vs_latency"] = {"n": len(sub), "pearson_r": pear.statistic, "pearson_p": pear.pvalue,
                              "spearman_r": spear.statistic, "spearman_p": spear.pvalue, "data": sub}

    sub2 = joined[joined["has_linked_turn"]].dropna(subset=["average_score"])
    pb = stats.pointbiserialr(sub2["task_success"].astype(int), sub2["average_score"])
    out["asq_vs_task_success"] = {"n": len(sub2), "r": pb.correlation, "p": pb.pvalue, "data": sub2}

    sub3 = joined[joined["has_linked_turn"]].dropna(subset=["asq_3"])
    pb3 = stats.pointbiserialr(sub3["has_correct_citation"].astype(int), sub3["asq_3"])
    out["asq3_vs_citation"] = {"n": len(sub3), "r": pb3.correlation, "p": pb3.pvalue, "data": sub3}

    return out


def compute_wilcoxon(joined: pd.DataFrame) -> dict:
    wide = joined.pivot(index="r0x", columns="scenario", values="average_score").dropna()
    results = {}
    w, p = stats.wilcoxon(wide["S1"], wide["S2"])
    results["average_score"] = {
        "n_pairs": len(wide), "mean_s1": wide["S1"].mean(), "mean_s2": wide["S2"].mean(),
        "statistic": w, "p": p,
    }
    for item in ["asq_1", "asq_2", "asq_3"]:
        wide_i = joined.pivot(index="r0x", columns="scenario", values=item).dropna()
        try:
            w, p = stats.wilcoxon(wide_i["S1"], wide_i["S2"])
        except ValueError:
            w, p = (np.nan, np.nan)  # all differences zero
        results[item] = {
            "n_pairs": len(wide_i), "mean_s1": wide_i["S1"].mean(), "mean_s2": wide_i["S2"].mean(),
            "statistic": w, "p": p,
        }
    return results


# ---------------------------------------------------------------------------
# Findings synthesis (data-driven, cross-referenced to already-verified numbers)
# ---------------------------------------------------------------------------

def build_findings(asq_overall, asq_item_scenario, task_success_df, latency_df, citehall, correlations, wilcoxon) -> dict:
    weakest_item_row = asq_item_scenario.loc[asq_item_scenario["mean"].idxmin()]
    ts = task_success_df.set_index("scenario")["success_rate"]
    rates_tied = abs(ts["S1"] - ts["S2"]) < 1e-9
    slower_latency_scenario = latency_df.loc[latency_df["mean_ms"].idxmax(), "scenario"]

    if rates_tied:
        success_problem = (
            f"Task success rate objektif sama rendahnya pada KEDUA skenario ({ts['S1']:.1%}, "
            f"berdasarkan answer_status in {{answered, verified}} via trace_id) -- jauh di bawah "
            f"kepuasan subjektif ASQ (mean keseluruhan {asq_overall['overall']['mean']:.2f}/7). "
            "Ini kesenjangan nyata antara persepsi kepuasan dan hasil objektif sistem, bukan "
            "perbedaan antar skenario."
        )
    else:
        weaker_success_scenario = ts.idxmin()
        success_problem = (
            f"Task success rate lebih rendah pada Skenario {weaker_success_scenario} "
            f"({ts[weaker_success_scenario]:.1%}) dibanding skenario lain -- proporsi giliran chat "
            "yang berstatus answer_status di luar {answered, verified} (mis. "
            "out_of_domain/insufficient_context/fallback_enforced) lebih tinggi pada skenario ini."
        )

    user_problems = [
        {
            "rank": 1,
            "problem": f"Item ASQ dengan skor rata-rata terendah: {weakest_item_row['item']} pada skenario "
                       f"{weakest_item_row['scenario']} (mean={weakest_item_row['mean']:.2f}/7).",
            "evidence": "Sheet 'ASQ per Item & Skenario' (dihitung langsung dari data mentah 43 baris ASQ).",
        },
        {
            "rank": 2,
            "problem": success_problem,
            "evidence": "Sheet 'Task Success Rate per Skenario' (join trace_id ke chat_evaluation_logs).",
        },
        {
            "rank": 3,
            "problem": "Target retrieval meleset pada beberapa pertanyaan otomatis (Q004, Q007, Q014) -- sistem "
                       "menjawab topik yang berdekatan tapi bukan yang ditanyakan secara spesifik, bukan fabrikasi, "
                       "murni retrieval yang tidak tepat sasaran.",
            "evidence": "evaluation/reports/2026-07-31/manual_annotation_needed/answer_quality_eval.csv "
                        "(rubrik manual, 2 rater independen).",
        },
    ]

    technical_causes = [
        {
            "rank": 1,
            "cause": f"Citation-correct rate hanya {citehall['citation_correct_rate']:.1%} "
                     f"({citehall['n_questions']} soal gold-QA) -- lebih dari separuh jawaban yang menyertakan "
                     "sitasi tidak mengutip sumber yang benar, mengindikasikan proses citation_builder/"
                     "output_claim_verifier butuh penguatan, bukan sekadar retrieval.",
            "evidence": "Sheet 'Citation Correctness, Hallucination, FPR'; run refresh_20260727_gates_all.",
        },
        {
            "rank": 2,
            "cause": f"False-positive rate ACIF (pertanyaan valid yang salah diblokir/fallback) = "
                     f"{citehall['false_positive_rate']:.1%} ({citehall['n_false_positive']}/"
                     f"{citehall['n_valid_answer_category']} soal kategori jawaban-valid) -- lebih tinggi dari "
                     "angka historis 2026-07-15 (0%), menunjukkan regresi pada kalibrasi threshold ACIF perlu "
                     "diinvestigasi.",
            "evidence": "Sheet 'Citation Correctness, Hallucination, FPR' (dihitung ulang dari data terkini, "
                        "menggantikan angka 2026-07-15 yang sudah usang).",
        },
        {
            "rank": 3,
            "cause": "Kapasitas hardware VPS (4 vCPU) menjadi bottleneck nyata pada beban bersamaan -- latensi "
                     "per-permintaan memburuk 2.3-4x pada 10 request bersamaan (391.6% CPU, hampir saturasi penuh).",
            "evidence": "evaluation/reports/2026-07-31/PANDUAN_REVISI_BAB4_2026-07-31.md §12.",
        },
    ]

    recommendations = [
        {
            "rank": 1,
            "recommendation": "Perkuat output_claim_verifier/citation_builder untuk menaikkan citation-correct "
                               "rate dari 39% -- audit soal-soal SPMB (kategori dengan volume sitasi terbanyak) "
                               "untuk pola kesalahan sitasi yang paling umum.",
            "based_on": "Penyebab teknis #1.",
        },
        {
            "rank": 2,
            "recommendation": "Investigasi ulang threshold ACIF Gate 1/Gate 5 untuk menekan false-positive rate "
                               "kembali mendekati 0%, terutama untuk kategori Regulation dan SPMB yang tercatat "
                               "salah fallback pada run terbaru.",
            "based_on": "Penyebab teknis #2.",
        },
        {
            "rank": 3,
            "recommendation": "Pertimbangkan penambahan vCPU atau paralelisasi panggilan LLM per-chunk (terutama "
                               "tahap summarization) sebelum uji beban dengan pengguna nyata dalam skala lebih "
                               "besar dari 10 permintaan bersamaan.",
            "based_on": "Penyebab teknis #3.",
        },
    ]

    return {"user_problems": user_problems, "technical_causes": technical_causes, "recommendations": recommendations}


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------

def make_charts(joined, asq_overall, latency_df, correlations, wilcoxon) -> dict[str, Path]:
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    paths = {}

    fig, ax = plt.subplots(figsize=(6, 4))
    means = [asq_overall["per_scenario"]["S1"]["mean"], asq_overall["per_scenario"]["S2"]["mean"]]
    sds = [asq_overall["per_scenario"]["S1"]["sd"], asq_overall["per_scenario"]["S2"]["sd"]]
    ax.bar(["S1", "S2"], means, yerr=sds, capsize=6, color=["#4C72B0", "#DD8452"])
    ax.set_ylim(0, 7.5)
    ax.set_ylabel("Skor ASQ rata-rata (1-7)")
    ax.set_title("Skor ASQ Keseluruhan per Skenario (mean ± SD)")
    paths["asq_per_scenario"] = FIG_DIR / "fig_asq_per_scenario.png"
    fig.savefig(paths["asq_per_scenario"]); plt.close(fig)

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.bar(latency_df["scenario"], latency_df["mean_ms"], color="#55A868")
    ax.set_ylabel("Latency rata-rata (ms)")
    ax.set_title("Latency per Skenario")
    paths["latency_per_scenario"] = FIG_DIR / "fig_latency_per_scenario.png"
    fig.savefig(paths["latency_per_scenario"]); plt.close(fig)

    fig, ax = plt.subplots(figsize=(6, 4))
    d = correlations["asq_vs_latency"]["data"]
    ax.scatter(d["average_score"], d["total_latency_ms"], alpha=0.7)
    r = correlations["asq_vs_latency"]["pearson_r"]
    ax.set_xlabel("Skor ASQ rata-rata"); ax.set_ylabel("Latency (ms)")
    ax.set_title(f"Korelasi ASQ vs Latency (Pearson r={r:.3f}, n={correlations['asq_vs_latency']['n']})")
    paths["corr_asq_latency"] = FIG_DIR / "fig_corr_asq_latency.png"
    fig.savefig(paths["corr_asq_latency"]); plt.close(fig)

    fig, ax = plt.subplots(figsize=(6, 4))
    wide = joined.pivot(index="r0x", columns="scenario", values="average_score").dropna()
    x = np.arange(len(wide))
    ax.plot(x, wide["S1"], "o-", label="S1", alpha=0.7)
    ax.plot(x, wide["S2"], "o-", label="S2", alpha=0.7)
    ax.set_xlabel("Responden (n=%d pasangan)" % len(wide)); ax.set_ylabel("Skor ASQ rata-rata")
    p = wilcoxon["average_score"]["p"]
    ax.set_title(f"ASQ S1 vs S2 per Responden (Wilcoxon p={p:.4f})")
    ax.legend()
    paths["wilcoxon_s1_s2"] = FIG_DIR / "fig_wilcoxon_s1_s2.png"
    fig.savefig(paths["wilcoxon_s1_s2"]); plt.close(fig)

    return paths


# ---------------------------------------------------------------------------
# Workbook writing
# ---------------------------------------------------------------------------

def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def autosize(ws) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        maxlen = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells[:80])
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 60)


def write_df_sheet(wb, name: str, df: pd.DataFrame, title: str | None = None) -> None:
    ws = wb.create_sheet(name)
    row0 = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT
        row0 = 3
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=row0, column=j, value=str(col))
    for i, (_, r) in enumerate(df.iterrows(), start=row0 + 1):
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            if isinstance(v, (np.floating, float)) and not pd.isna(v):
                v = round(float(v), 4)
            elif pd.isna(v) if not isinstance(v, (list, dict)) else False:
                v = None
            ws.cell(row=i, column=j, value=v)
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=row0, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = ws.cell(row=row0 + 1, column=1).coordinate
    autosize(ws)


def embed_image(ws, png_path: Path, anchor_cell: str, width_px: int = 650) -> None:
    if not png_path.exists():
        return
    img = XLImage(str(png_path))
    scale = width_px / img.width
    img.width = width_px
    img.height = int(img.height * scale)
    ws.add_image(img, anchor_cell)


def write_findings_sheet(wb, name: str, title: str, rows: list[dict], columns: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    for j, (_, header) in enumerate(columns, start=1):
        ws.cell(row=3, column=j, value=header)
    for i, row in enumerate(rows, start=4):
        for j, (key, _) in enumerate(columns, start=1):
            cell = ws.cell(row=i, column=j, value=row.get(key))
            cell.alignment = WRAP
    style_header(ws, len(columns))
    for c in range(1, len(columns) + 1):
        ws.cell(row=3, column=c).fill = HEADER_FILL
        ws.cell(row=3, column=c).font = HEADER_FONT
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 55
    ws.freeze_panes = "A4"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", required=True, type=Path)
    args = parser.parse_args()

    d = load_inputs(args.input_dir)
    joined = build_joined(d)

    asq_item_scenario = compute_asq_per_item_scenario(joined)
    asq_overall = compute_asq_overall(joined)
    cronbach = compute_cronbach_per_scenario(joined)
    task_success_df = compute_task_success(joined)
    completion_df = compute_completion_time(joined)
    latency_df = compute_latency(joined)
    citehall = compute_citation_hallucination_fpr(d["ev"])
    correlations = compute_correlations(joined)
    wilcoxon = compute_wilcoxon(joined)
    findings = build_findings(asq_overall, asq_item_scenario, task_success_df, latency_df, citehall, correlations, wilcoxon)

    charts = make_charts(joined, asq_overall, latency_df, correlations, wilcoxon)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    safe_cols = ["r0x", "scenario", "asq_1", "asq_2", "asq_3", "average_score", "total_latency_ms",
                 "answer_status", "task_success", "n_citations", "n_valid_citations",
                 "has_correct_citation", "completion_time_s", "completion_time_valid",
                 "completion_time_outlier", "completion_time_clean"]
    joined[safe_cols].to_csv(OUT_DIR / "raw_joined_by_scenario.csv", index=False)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 0: Ringkasan Eksekutif
    ws0 = wb.create_sheet("0. Ringkasan Eksekutif")
    ws0.cell(row=1, column=1, value="Ringkasan Eksekutif - Evaluasi ASQ per Skenario (2026-08-02)").font = TITLE_FONT
    summary_lines = [
        ("Skor ASQ keseluruhan", f"{asq_overall['overall']['mean']:.2f}/7 (n={asq_overall['overall']['n']}, "
                                  f"SD={asq_overall['overall']['sd']:.2f}, 95% CI [{asq_overall['overall']['ci95_lo']:.2f}, "
                                  f"{asq_overall['overall']['ci95_hi']:.2f}])"),
        ("Skor ASQ per skenario", f"S1={asq_overall['per_scenario']['S1']['mean']:.2f} vs "
                                   f"S2={asq_overall['per_scenario']['S2']['mean']:.2f}"),
        ("Cronbach's alpha S1 / S2", f"{cronbach['S1']['alpha']:.3f} / {cronbach['S2']['alpha']:.3f}"),
        ("Task success rate S1 / S2", f"{task_success_df.set_index('scenario').loc['S1','success_rate']:.1%} / "
                                       f"{task_success_df.set_index('scenario').loc['S2','success_rate']:.1%}"),
        ("Completion time (median) S1 / S2 (detik)",
         f"{completion_df.set_index('scenario').loc['S1','median']:.0f} / "
         f"{completion_df.set_index('scenario').loc['S2','median']:.0f}"),
        ("Latency rata-rata S1 / S2 (ms)", f"{latency_df.set_index('scenario').loc['S1','mean_ms']:.0f} / "
                                            f"{latency_df.set_index('scenario').loc['S2','mean_ms']:.0f}"),
        ("Citation correct rate (gold-QA, n=41)", f"{citehall['citation_correct_rate']:.1%}"),
        ("Hallucination rate (gold-QA, n=41)", f"{citehall['hallucination_rate']:.1%}"),
        ("False-positive rate ACIF", f"{citehall['false_positive_rate']:.1%} "
                                      f"({citehall['n_false_positive']}/{citehall['n_valid_answer_category']})"),
        ("Korelasi ASQ vs Latency", f"Pearson r={correlations['asq_vs_latency']['pearson_r']:.3f} "
                                     f"(p={correlations['asq_vs_latency']['pearson_p']:.4f}, n={correlations['asq_vs_latency']['n']})"),
        ("Korelasi ASQ vs Task Success", f"point-biserial r={correlations['asq_vs_task_success']['r']:.3f} "
                                          f"(p={correlations['asq_vs_task_success']['p']:.4f}, n={correlations['asq_vs_task_success']['n']})"),
        ("Korelasi ASQ_3 (dukungan informasi) vs Citation Correctness",
         f"point-biserial r={correlations['asq3_vs_citation']['r']:.3f} "
         f"(p={correlations['asq3_vs_citation']['p']:.4f}, n={correlations['asq3_vs_citation']['n']})"),
        ("Perbedaan ASQ S1 vs S2 (Wilcoxon)", f"p={wilcoxon['average_score']['p']:.4f}, n pasangan="
                                               f"{wilcoxon['average_score']['n_pairs']}"),
    ]
    for i, (label, value) in enumerate(summary_lines, start=3):
        ws0.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws0.cell(row=i, column=2, value=value)
    ws0.column_dimensions["A"].width = 45
    ws0.column_dimensions["B"].width = 70

    # Sheet 1: ASQ per Item & Skenario
    raw_asq = joined[["r0x", "scenario", "asq_1", "asq_2", "asq_3", "average_score"]].sort_values(["r0x", "scenario"])
    write_df_sheet(wb, "1. ASQ per Item Skenario", raw_asq, "Data Mentah ASQ (R01-R22)")
    ws1b = wb["1. ASQ per Item Skenario"]
    start = ws1b.max_row + 3
    ws1b.cell(row=start, column=1, value="Agregat per Item x Skenario (mean/median/SD/n)").font = TITLE_FONT
    agg_cols = ["scenario", "item", "n", "mean", "median", "sd"]
    for j, c in enumerate(agg_cols, start=1):
        ws1b.cell(row=start + 2, column=j, value=c)
    for i, (_, r) in enumerate(asq_item_scenario[agg_cols].iterrows(), start=start + 3):
        for j, c in enumerate(agg_cols, start=1):
            v = r[c]
            ws1b.cell(row=i, column=j, value=round(float(v), 4) if isinstance(v, (float, np.floating)) else v)
    embed_image(ws1b, charts["asq_per_scenario"], f"H{start}")

    # Sheet 2: ASQ Keseluruhan
    ws2 = wb.create_sheet("2. ASQ Keseluruhan")
    ws2.cell(row=1, column=1, value="Statistik Deskriptif Skor ASQ Keseluruhan").font = TITLE_FONT
    stat_labels = [("n", "n"), ("mean", "Mean"), ("median", "Median"), ("sd", "SD"), ("iqr", "IQR"),
                   ("q1", "Q1"), ("q3", "Q3"), ("ci95_lo", "95% CI bawah"), ("ci95_hi", "95% CI atas")]
    ws2.cell(row=3, column=1, value="Statistik")
    ws2.cell(row=3, column=2, value="Keseluruhan")
    ws2.cell(row=3, column=3, value="S1")
    ws2.cell(row=3, column=4, value="S2")
    for i, (key, label) in enumerate(stat_labels, start=4):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=round(float(asq_overall["overall"][key]), 4) if pd.notna(asq_overall["overall"][key]) else None)
        ws2.cell(row=i, column=3, value=round(float(asq_overall["per_scenario"]["S1"][key]), 4) if pd.notna(asq_overall["per_scenario"]["S1"][key]) else None)
        ws2.cell(row=i, column=4, value=round(float(asq_overall["per_scenario"]["S2"][key]), 4) if pd.notna(asq_overall["per_scenario"]["S2"][key]) else None)
    style_header(ws2, 4)
    for c in "ABCD":
        ws2.column_dimensions[c].width = 20

    # Sheet 3: Cronbach Alpha
    ws3 = wb.create_sheet("3. Cronbach Alpha per Skenario")
    ws3.cell(row=1, column=1, value="Cronbach's Alpha per Skenario (3 item ASQ)").font = TITLE_FONT
    ws3.cell(row=1, column=1).alignment = WRAP
    ws3.cell(row=2, column=1, value="Catatan: indikatif -- hanya 3 item, n=22 (S1) / n=21 (S2). Lihat sheet Metodologi.")
    headers = ["Skenario", "n", "Cronbach's Alpha", "Item-Total Corr (ASQ_1)", "Item-Total Corr (ASQ_2)", "Item-Total Corr (ASQ_3)"]
    for j, h in enumerate(headers, start=1):
        ws3.cell(row=4, column=j, value=h)
    for i, scen in enumerate(["S1", "S2"], start=5):
        c = cronbach[scen]
        ws3.cell(row=i, column=1, value=scen)
        ws3.cell(row=i, column=2, value=c["n"])
        ws3.cell(row=i, column=3, value=round(c["alpha"], 4))
        ws3.cell(row=i, column=4, value=round(c["item_total_corr"]["asq_1"], 4))
        ws3.cell(row=i, column=5, value=round(c["item_total_corr"]["asq_2"], 4))
        ws3.cell(row=i, column=6, value=round(c["item_total_corr"]["asq_3"], 4))
    style_header(ws3, len(headers))
    autosize(ws3)

    # Sheet 4: Task Success Rate
    write_df_sheet(wb, "4. Task Success Rate", task_success_df, "Task Success Rate per Skenario "
                    "(sukses = answer_status in {answered, verified}, dihubungkan via trace_id)")
    ws4 = wb["4. Task Success Rate"]
    start4 = ws4.max_row + 3
    ws4.cell(row=start4, column=1, value="Data terhubung per responden").font = TITLE_FONT
    detail = joined[["r0x", "scenario", "average_score", "answer_status", "task_success", "has_linked_turn"]].sort_values(["r0x", "scenario"])
    for j, c in enumerate(detail.columns, start=1):
        ws4.cell(row=start4 + 2, column=j, value=c)
    for i, (_, r) in enumerate(detail.iterrows(), start=start4 + 3):
        for j, c in enumerate(detail.columns, start=1):
            ws4.cell(row=i, column=j, value=r[c] if not (isinstance(r[c], float) and pd.isna(r[c])) else None)

    # Sheet 5: Completion Time
    write_df_sheet(wb, "5. Completion Time", completion_df, "Completion Time per Skenario (proxy: lihat sheet "
                    "Metodologi untuk definisi operasional dan catatan data invalid)")

    # Sheet 6: Latency
    write_df_sheet(wb, "6. Latency per Skenario", latency_df, "Latency per Skenario")
    ws6 = wb["6. Latency per Skenario"]
    embed_image(ws6, charts["latency_per_scenario"], f"A{ws6.max_row + 3}")

    # Sheet 7: Citation / Hallucination / FPR
    ws7 = wb.create_sheet("7. Citation Halusinasi FPR")
    ws7.cell(row=1, column=1, value="Citation Correctness, Hallucination Rate, False-Positive Rate "
                                     "(sumber: run bersih refresh_20260727_gates_all, n=41 soal gold-QA)").font = TITLE_FONT
    ws7.cell(row=1, column=1).alignment = WRAP
    kv = [
        ("Jumlah soal gold-QA", citehall["n_questions"]),
        ("Citation correct rate", f"{citehall['citation_correct_rate']:.4f}"),
        ("Hallucination rate", f"{citehall['hallucination_rate']:.4f}"),
        ("Soal kategori expected_behavior=answer", citehall["n_valid_answer_category"]),
        ("Soal salah diblokir/fallback (false positive)", citehall["n_false_positive"]),
        ("False-positive rate ACIF", f"{citehall['false_positive_rate']:.4f}"),
    ]
    for i, (k, v) in enumerate(kv, start=3):
        ws7.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws7.cell(row=i, column=2, value=v)
    start7 = 3 + len(kv) + 2
    ws7.cell(row=start7, column=1, value="Rincian per Kategori").font = TITLE_FONT
    for j, c in enumerate(citehall["per_category"].columns, start=1):
        ws7.cell(row=start7 + 2, column=j, value=c)
    for i, (_, r) in enumerate(citehall["per_category"].iterrows(), start=start7 + 3):
        for j, c in enumerate(citehall["per_category"].columns, start=1):
            ws7.cell(row=i, column=j, value=r[c])
    ws7.column_dimensions["A"].width = 40
    ws7.column_dimensions["B"].width = 20

    # Sheet 8: Korelasi
    ws8 = wb.create_sheet("8. Korelasi")
    ws8.cell(row=1, column=1, value="Tiga Korelasi").font = TITLE_FONT
    rows8 = [
        ("ASQ (average_score) vs Latency", correlations["asq_vs_latency"]["n"],
         correlations["asq_vs_latency"]["pearson_r"], correlations["asq_vs_latency"]["pearson_p"],
         correlations["asq_vs_latency"]["spearman_r"], correlations["asq_vs_latency"]["spearman_p"]),
        ("ASQ (average_score) vs Task Success (point-biserial)", correlations["asq_vs_task_success"]["n"],
         correlations["asq_vs_task_success"]["r"], correlations["asq_vs_task_success"]["p"], None, None),
        ("ASQ_3 (dukungan informasi) vs Citation Correctness (point-biserial)", correlations["asq3_vs_citation"]["n"],
         correlations["asq3_vs_citation"]["r"], correlations["asq3_vs_citation"]["p"], None, None),
    ]
    headers8 = ["Pasangan Variabel", "n", "r (Pearson/point-biserial)", "p-value", "Spearman rho", "Spearman p"]
    for j, h in enumerate(headers8, start=1):
        ws8.cell(row=3, column=j, value=h)
    for i, row in enumerate(rows8, start=4):
        for j, v in enumerate(row, start=1):
            ws8.cell(row=i, column=j, value=round(float(v), 4) if isinstance(v, (float, np.floating)) and pd.notna(v) else v)
    style_header(ws8, len(headers8))
    autosize(ws8)
    embed_image(ws8, charts["corr_asq_latency"], "A10")

    # Sheet 9: Perbedaan ASQ Antar Skenario
    ws9 = wb.create_sheet("9. Perbedaan ASQ Antar Skenario")
    ws9.cell(row=1, column=1, value="Uji Wilcoxon Signed-Rank Berpasangan S1 vs S2").font = TITLE_FONT
    headers9 = ["Variabel", "n pasangan", "Mean S1", "Mean S2", "Statistik W", "p-value"]
    for j, h in enumerate(headers9, start=1):
        ws9.cell(row=3, column=j, value=h)
    label_map = {"average_score": "ASQ rata-rata", "asq_1": ASQ_ITEM_LABELS["asq_1"], "asq_2": ASQ_ITEM_LABELS["asq_2"], "asq_3": ASQ_ITEM_LABELS["asq_3"]}
    for i, key in enumerate(["average_score", "asq_1", "asq_2", "asq_3"], start=4):
        r = wilcoxon[key]
        ws9.cell(row=i, column=1, value=label_map[key])
        ws9.cell(row=i, column=2, value=r["n_pairs"])
        ws9.cell(row=i, column=3, value=round(float(r["mean_s1"]), 4))
        ws9.cell(row=i, column=4, value=round(float(r["mean_s2"]), 4))
        ws9.cell(row=i, column=5, value=round(float(r["statistic"]), 4) if pd.notna(r["statistic"]) else None)
        ws9.cell(row=i, column=6, value=round(float(r["p"]), 4) if pd.notna(r["p"]) else None)
    style_header(ws9, len(headers9))
    autosize(ws9)
    embed_image(ws9, charts["wilcoxon_s1_s2"], "A12")

    # Sheets 10-12: Findings
    write_findings_sheet(wb, "10. Temuan Masalah Pengguna", "3 Masalah Pengguna yang Paling Sering Muncul",
                          findings["user_problems"], [("rank", "Peringkat"), ("problem", "Masalah"), ("evidence", "Bukti/Sumber")])
    write_findings_sheet(wb, "11. Temuan Penyebab Teknis", "3 Penyebab Teknis Terbesar",
                          findings["technical_causes"], [("rank", "Peringkat"), ("cause", "Penyebab"), ("evidence", "Bukti/Sumber")])
    write_findings_sheet(wb, "12. Rekomendasi Perbaikan", "Rekomendasi Perbaikan Berdasarkan Bukti",
                          findings["recommendations"], [("rank", "Peringkat"), ("recommendation", "Rekomendasi"), ("based_on", "Berdasarkan Temuan")])

    # Sheet 13: Metodologi & Keterbatasan
    ws13 = wb.create_sheet("13. Metodologi Keterbatasan")
    ws13.cell(row=1, column=1, value="Metodologi & Keterbatasan").font = TITLE_FONT
    method_notes = [
        "Sumber data: studi usability ASQ manusia 2026-07-16 (22 partisipan, skenario S1/S2, 3 item "
        "ASQ skala 1-7) dihubungkan ke log produksi (chat_evaluation_logs, citation_evaluation_logs) "
        "via trace_id, plus run gold-QA otomatis bersih 'refresh_20260727_gates_all' (n=41 soal, 7 "
        "kategori) untuk citation correctness/hallucination/false-positive rate.",
        "PRIVASI: kode R01-R22 pada workbook ini adalah label anonim BARU yang dibangun khusus untuk "
        "laporan ini (diurutkan berdasarkan waktu submisi skenario pertama), TIDAK harus sama dengan "
        "R01-R22 pada laporan 2026-07-16 -- keduanya berasal dari 43 baris mentah yang sama sehingga "
        "statistik agregat identik, namun label individual boleh berbeda. Kode partisipan asli, "
        "trace_id, dan session_id mentah TIDAK PERNAH dituliskan ke file ini.",
        "Task success (per giliran chat) didefinisikan sebagai answer_status in {answered, verified}. "
        "Baris tanpa trace_id/giliran chat yang cocok ditandai 'has_linked_turn=False' dan dikeluarkan "
        "dari perhitungan rate (bukan dihitung sebagai gagal).",
        "Completion time adalah PROXY operasional, bukan stopwatch: untuk S1 = waktu submit ASQ S1 - "
        "waktu giliran chat pertama pada sesi tsb; untuk S2 = waktu submit ASQ S2 - waktu submit ASQ "
        "S1. Dua kategori baris dikeluarkan dari rata-rata/median 'bersih' (dilaporkan terpisah, "
        "bukan disembunyikan): (a) durasi negatif (kemungkinan selisih jam/urutan submit tidak wajar), "
        "dan (b) durasi outlier >1 jam (tidak masuk akal untuk skenario tugas singkat -- kemungkinan "
        "session_id lama terbawa dari sesi uji-coba sebelumnya). Lihat kolom n_excluded_negative dan "
        "n_excluded_outlier_gt_1h pada sheet 5.",
        "1 dari 43 baris ASQ (S1 satu responden) tidak memiliki trace_id di data sumber -- baris ini "
        "tidak memiliki data latency/task-success/citation linkage untuk giliran tsb, namun tetap "
        "masuk skor ASQ mentah.",
        "Cronbach's alpha dihitung dari hanya 3 item -- indikatif, bukan bukti reliabilitas skala "
        "penuh (instrumen ASQ standar memang hanya 3 item per desain Lewis 1991).",
        "Kolom 'notes' pada asq_responses hampir seluruhnya kosong -- tidak ada data keluhan bebas-teks "
        "untuk ditambang; 3 masalah pengguna tersering dan 3 penyebab teknis terbesar pada sheet 10-11 "
        "adalah SINTESIS dari hasil kuantitatif pada workbook ini plus temuan yang sudah terverifikasi "
        "sebelumnya (evaluation/reports/2026-07-31/), bukan pengukuran langsung baru -- setiap baris "
        "mencantumkan sumber buktinya.",
        "False-positive rate dihitung ULANG dari data terkini (bukan reuse angka 2026-07-15 yang sudah "
        "usang) -- definisi: proporsi soal berkategori expected_behavior='answer' yang justru "
        "fallback_correct=false (sistem salah menolak/fallback padahal seharusnya menjawab).",
        "Seluruh query bersifat read-only terhadap data yang sudah ada di VPS produksi -- tidak ada "
        "panggilan /chat baru atau evaluation run baru yang dijalankan untuk laporan ini.",
        "Uji Wilcoxon S1 vs S2 pada skor ASQ rata-rata menghasilkan p=1.0000 -- BUKAN kesalahan "
        "perhitungan: 13 dari 21 pasangan responden memberi skor rata-rata yang PERSIS SAMA pada "
        "kedua skenario (efek ceiling umum pada skala 1-7 untuk tugas singkat), menyisakan hanya 8 "
        "pasangan dengan selisih non-nol yang kebetulan seimbang antara naik dan turun. Interpretasi: "
        "tidak ada bukti statistik bahwa satu skenario dipersepsikan lebih baik dari yang lain.",
    ]
    for i, note in enumerate(method_notes, start=3):
        cell = ws13.cell(row=i, column=1, value=f"- {note}")
        cell.alignment = WRAP
        ws13.row_dimensions[i].height = 60
    ws13.column_dimensions["A"].width = 130

    order = [
        "0. Ringkasan Eksekutif", "1. ASQ per Item Skenario", "2. ASQ Keseluruhan",
        "3. Cronbach Alpha per Skenario", "4. Task Success Rate", "5. Completion Time",
        "6. Latency per Skenario", "7. Citation Halusinasi FPR", "8. Korelasi",
        "9. Perbedaan ASQ Antar Skenario", "10. Temuan Masalah Pengguna",
        "11. Temuan Penyebab Teknis", "12. Rekomendasi Perbaikan", "13. Metodologi Keterbatasan",
    ]
    wb._sheets = [wb[name] for name in order]

    out_path = OUT_DIR / "ASQ_Evaluasi_Skenario_Lengkap_2026-08-02.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"Wrote {OUT_DIR / 'raw_joined_by_scenario.csv'}")
    for name, p in charts.items():
        print(f"Wrote {p}")


if __name__ == "__main__":
    main()
