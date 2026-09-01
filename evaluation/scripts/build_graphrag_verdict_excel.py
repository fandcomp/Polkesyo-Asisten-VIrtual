"""GraphRAG isolation experiment (2026-07-25): full evaluation report + verdict workbook.

Compares Vector-RAG-only vs Vector-RAG+GraphRAG (ACIF held constant at gates_all) across 3
segments: the 6 new multi-hop probe questions (decisive evidence), the 14 existing SPMB
questions (graph-triggering but mostly single-fact), and everything else (sanity-check control,
should show ~no difference).

Produces:
- Sheet 0: Ringkasan (plain-language verdict + caveats)
- Sheet 1: Daftar Soal & Kondisi Uji (test catalog -- all 47 unique questions x conditions tested)
- Sheet 2: Hasil Uji per Soal per Kondisi (94 rows: question, expected answer, ACTUAL system
  answer, every scored metric, per condition)
- Sheet 3: Ringkasan Segmen (3-segment aggregate comparison table)
- Sheet 4: Grafik per Aspek (8 separate charts, one per evaluated aspect, not one combined chart)
- Sheet 5: Data Mentah (raw per-question data, audit trail)
"""
from __future__ import annotations

from pathlib import Path

import matplotlib
import openpyxl
import pandas as pd

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).parent.parent.parent
# 2026-07-26 = refreshed data after re-ingesting 3 documents (Reguler SMA/RPL/Profesi) with the
# fixed chunk_text_structured section-boundary chunking and re-pinned gold-QA/multihop ground
# truth chunk IDs -- supersedes 2026-07-25c, which scored against the pre-fix chunks. The old
# data is kept on disk for audit trail, not deleted.
RAW = ROOT / "evaluation/reports/2026-07-26/raw_exports"
OUT_DIR = ROOT / "evaluation/reports/2026-07-26"
FIG_DIR = OUT_DIR / "figures"

GATE_RAW = RAW
GATE_FIG_DIR = FIG_DIR

GATE_LABELS = {
    "gates_none": "0 Gate Aktif (Semua Nonaktif)",
    "gates_1": "Gate 1 Saja",
    "gates_1_2": "Gate 1-2",
    "gates_1_2_3": "Gate 1-3",
    "gates_1_2_3_4": "Gate 1-4",
    "gates_all_minus_1": "Semua kecuali Gate 1",
    "gates_all_minus_2": "Semua kecuali Gate 2",
    "gates_all_minus_3": "Semua kecuali Gate 3",
    "gates_all_minus_4": "Semua kecuali Gate 4",
    "gates_all_minus_5": "Semua kecuali Gate 5 (≡ Gate 1-4)",
    "gates_all": "5 Gate Aktif (Semua)",
}
GATE_ASPECTS = [
    ("precision_at_3", "Precision@3"), ("recall_at_3", "Recall@3"),
    ("precision_at_5", "Precision@5"), ("recall_at_5", "Recall@5"), ("hit_rate_at_5", "Hit Rate@5"),
    ("retrieval_relevance_score", "Retrieval Relevance (LLM-Judge)"),
    ("citation_coverage", "Citation Coverage"), ("citation_correctness", "Citation Correctness"),
    ("fallback_correctness", "Fallback Correctness"), ("attack_success_rate", "Attack Success Rate"),
    ("faithfulness_score", "Faithfulness"), ("answer_relevance_score", "Answer Relevance"),
    ("average_latency_ms", "Latency Rata-rata (ms)"),
]
GATE_CONFIG_NAMES = [
    "gates_none", "gates_1", "gates_1_2", "gates_1_2_3", "gates_1_2_3_4",
    "gates_all_minus_1", "gates_all_minus_2", "gates_all_minus_3", "gates_all_minus_4", "gates_all",
]

HEADER_FILL = PatternFill(start_color="FF1F3B3C", end_color="FF1F3B3C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

SPMB_QIDS = {f"Q{n:03d}" for n in range(1, 15)}  # Q001-Q014

RUN_NAME_TO_CONDITION = {
    "refresh_20260726_41q_graph_on": ("41q", "graph_on"),
    "refresh_20260726_41q_graph_off": ("41q", "graph_off"),
    "refresh_20260726_mh_graph_on": ("mh", "graph_on"),
    "refresh_20260726_mh_graph_off": ("mh", "graph_off"),
}

SEGMENT_LABELS = {
    "1_multihop": "Segmen 1: Multi-Hop Baru (6 soal)",
    "2_spmb": "Segmen 2: SPMB Existing (14 soal)",
    "3_control": "Segmen 3: Non-SPMB / Kontrol (27 soal)",
}

CONDITION_LABELS = {"graph_on": "Graph ON (Vector+GraphRAG)", "graph_off": "Graph OFF (Vector RAG saja)"}

ASPECTS = [
    ("precision_at_3", "Precision@3", "nilai 0-1"),
    ("recall_at_3", "Recall@3", "nilai 0-1"),
    ("precision_at_5", "Precision@5", "nilai 0-1"),
    ("recall_at_5", "Recall@5", "nilai 0-1"),
    ("retrieval_relevance", "Retrieval Relevance (LLM-Judge)", "nilai 0-1"),
    ("citation_coverage_pct", "Citation Coverage", "%"),
    ("citation_correctness_pct", "Citation Correctness", "%"),
    ("fallback_correctness_pct", "Fallback Correctness", "%"),
    ("faithfulness", "Faithfulness (LLM-Judge)", "nilai 0-1"),
    ("answer_relevance", "Answer Relevance (LLM-Judge)", "nilai 0-1"),
    ("avg_latency_ms", "Latency Rata-rata", "ms"),
]


def load_gold_qa() -> pd.DataFrame:
    df = pd.read_json(ROOT / "backend/app/evaluation/gold_qa_dataset.jsonl", lines=True)
    df["dataset"] = "41q"
    return df


def load_multihop() -> pd.DataFrame:
    df = pd.read_json(ROOT / "backend/app/evaluation/graphrag_multihop_dataset.jsonl", lines=True)
    df["dataset"] = "mh"
    return df


def _load_all_refresh() -> pd.DataFrame:
    """Loads all 14 refreshed runs (10 gate configs + 4 GraphRAG conditions, 504 rows total) in
    one shot -- both axes now share one unified 2026-07-25c data export (post metric-fix)."""
    runs = pd.read_csv(RAW / "refresh_evaluation_runs.csv")[["id", "run_name", "config_name"]]
    results = pd.read_csv(RAW / "refresh_evaluation_results.csv")
    answers = pd.read_csv(RAW / "refresh_chat_answers.csv")[
        ["trace_id", "final_answer", "answer_status", "fallback_triggered", "fallback_reason", "model_used"]
    ]
    df = results.merge(runs, left_on="evaluation_run_id", right_on="id", suffixes=("", "_run"))
    df = df.merge(answers, on="trace_id", how="left")
    return df


def load_results() -> pd.DataFrame:
    """The 4 GraphRAG-condition rows (41q x on/off, mh x on/off = 94 rows) out of the unified export."""
    all_df = _load_all_refresh()
    df = all_df[all_df["run_name"].isin(RUN_NAME_TO_CONDITION.keys())].copy()
    df["dataset"] = df["run_name"].map(lambda r: RUN_NAME_TO_CONDITION[r][0])
    df["condition"] = df["run_name"].map(lambda r: RUN_NAME_TO_CONDITION[r][1])

    def segment(row):
        if row["dataset"] == "mh":
            return "1_multihop"
        return "2_spmb" if row["question_id"] in SPMB_QIDS else "3_control"

    df["segment"] = df.apply(segment, axis=1)
    return df


def load_gate_ablation_detail() -> pd.DataFrame:
    """The 10-gate-config rows (410 rows) out of the unified 2026-07-25c export -- post metric-fix,
    so precision_at_3/recall_at_3/precision_at_5/recall_at_5/retrieval_relevance_score are all now
    valid real numbers (unlike the first version of this report, built from pre-fix 2026-07-24 data)."""
    all_df = _load_all_refresh()
    df = all_df[all_df["config_name"].isin(GATE_CONFIG_NAMES) & all_df["run_name"].str.contains("gates_")].copy()
    df["gate_label"] = df["config_name"].map(lambda c: GATE_LABELS.get(c, c))
    return df


def load_gate_ablation_summary() -> pd.DataFrame:
    detail = load_gate_ablation_detail()
    rows = []
    for cfg in GATE_CONFIG_NAMES:
        sub = detail[detail["config_name"] == cfg]
        row = {"Kondisi Gate": GATE_LABELS.get(cfg, cfg), "Nama Config": cfg, "N Soal": len(sub)}
        citation_present = _to_bool_series(sub["citation_present"]).fillna(False).astype(bool)
        citation_correct = _to_bool_series(sub["citation_correct"]).fillna(False).astype(bool)
        fallback = _to_bool_series(sub["fallback_correct"]).dropna().astype(bool)
        attack = _to_bool_series(sub["attack_success"]).dropna().astype(bool)
        aspect_values = {
            "precision_at_3": _mean(sub["precision_at_3"]), "recall_at_3": _mean(sub["recall_at_3"]),
            "precision_at_5": _mean(sub["precision_at_5"]), "recall_at_5": _mean(sub["recall_at_5"]),
            "hit_rate_at_5": _mean(sub["hit_rate_at_5"]),
            "retrieval_relevance_score": _mean(sub["retrieval_relevance_score"]),
            "citation_coverage": round(citation_present.mean(), 4) if len(sub) else None,
            "citation_correctness": (
                round(citation_correct[citation_present].mean(), 4) if citation_present.any() else None
            ),
            "fallback_correctness": round(fallback.mean(), 4) if len(fallback) else None,
            "attack_success_rate": round(attack.mean(), 4) if len(attack) else None,
            "faithfulness_score": _mean(sub["faithfulness_score"]),
            "answer_relevance_score": _mean(sub["answer_relevance_score"]),
            "average_latency_ms": _mean(sub["total_latency_ms"]),
        }
        for key, label in GATE_ASPECTS:
            row[label] = aspect_values.get(key)
        rows.append(row)
    return pd.DataFrame(rows)


def _mean(s: pd.Series) -> float | None:
    s = s.dropna()
    return round(float(s.mean()), 4) if len(s) else None


_TRUE_STRINGS = {"t", "true", "1", "yes"}


def _to_bool_series(s: pd.Series) -> pd.Series:
    """Parse a boolean-ish column into real Python bools, NaN-preserving.

    `refresh_*.csv` is a `psql \\copy` export -- Postgres boolean columns come out as the
    literal text "t"/"f", not Python True/False. A bare `.astype(bool)` on that object-dtype
    column silently treats EVERY non-empty string as truthy (`bool("f") is True`), which is
    what made sheet 6's Citation Coverage / Fallback Correctness / Attack Success Rate columns
    read a flat 1.0 for all 10 gate configs -- contradicted by sheet 8 (built from a separate,
    correctly-typed CSV), which shows these metrics genuinely vary by gate. This mirrors the
    `r.get(...) in (True, "t")` check already used elsewhere in this file for the same reason.
    """
    def _one(v):
        if pd.isna(v):
            return None
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in _TRUE_STRINGS

    return s.map(_one)


def aggregate(df: pd.DataFrame, segment: str, condition: str) -> dict:
    sub = df[(df["segment"] == segment) & (df["condition"] == condition)]
    citation_present = _to_bool_series(sub["citation_present"]).fillna(False).astype(bool)
    citation_correct = _to_bool_series(sub["citation_correct"]).fillna(False).astype(bool)
    fallback = _to_bool_series(sub["fallback_correct"])
    return {
        "n": len(sub),
        "precision_at_3": _mean(sub["precision_at_3"]),
        "recall_at_3": _mean(sub["recall_at_3"]),
        "precision_at_5": _mean(sub["precision_at_5"]),
        "recall_at_5": _mean(sub["recall_at_5"]),
        "retrieval_relevance": _mean(sub["retrieval_relevance_score"]),
        "citation_coverage_pct": round(citation_present.mean() * 100, 1) if len(sub) else None,
        "citation_correctness_pct": (
            round(citation_correct[citation_present].mean() * 100, 1)
            if citation_present.any() else None
        ),
        "fallback_correctness_pct": (
            round(fallback.dropna().astype(bool).mean() * 100, 1) if fallback.notna().any() else None
        ),
        "faithfulness": _mean(sub["faithfulness_score"]),
        "answer_relevance": _mean(sub["answer_relevance_score"]),
        "avg_latency_ms": _mean(sub["total_latency_ms"]),
    }


def build_segment_table(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for seg in ["1_multihop", "2_spmb", "3_control"]:
        on = aggregate(df, seg, "graph_on")
        off = aggregate(df, seg, "graph_off")
        row = {"Segmen": SEGMENT_LABELS[seg], "N Soal": on["n"]}
        for key, label, _unit in ASPECTS:
            row[f"{label} - ON"] = on[key]
            row[f"{label} - OFF"] = off[key]
            row[f"{label} - Delta"] = (
                round(on[key] - off[key], 4) if on[key] is not None and off[key] is not None else None
            )
        rows.append(row)
    return pd.DataFrame(rows)


def make_charts(segment_table: pd.DataFrame) -> list[Path]:
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    labels = ["Multi-Hop\n(baru)", "SPMB\n(existing)", "Non-SPMB\n(kontrol)"]
    paths = []
    for key, label, unit in ASPECTS:
        on_col, off_col = f"{label} - ON", f"{label} - OFF"
        fig, ax = plt.subplots(figsize=(7, 4.5))
        x = range(len(labels))
        width = 0.35
        on_vals = segment_table[on_col].fillna(0)
        off_vals = segment_table[off_col].fillna(0)
        ax.bar([i - width / 2 for i in x], on_vals, width, label="Graph ON (Vector+GraphRAG)", color="#4C72B0")
        ax.bar([i + width / 2 for i in x], off_vals, width, label="Graph OFF (Vector RAG saja)", color="#C44E52")
        ax.set_xticks(list(x))
        ax.set_xticklabels(labels)
        ax.set_ylabel(unit)
        ax.set_title(f"GraphRAG Isolation: {label} per Segmen")
        ax.legend(fontsize=8)
        fname = f"fig_graphrag_{key}"
        path = FIG_DIR / f"{fname}.png"
        fig.savefig(path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        paths.append(path)
    return paths


def make_gate_charts(gate_summary: pd.DataFrame, gate_ranking: pd.DataFrame) -> list[Path]:
    """Fresh gate-axis charts from the 2026-07-25c refreshed data (supersedes the 2026-07-24
    pre-fix figures, which are kept on disk for audit trail but no longer embedded here)."""
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    labels = gate_summary["Kondisi Gate"].tolist()
    paths = []

    for key_label, title in [
        (["Precision@3", "Precision@5", "Recall@3", "Recall@5"], "Precision & Recall per Konfigurasi Gate"),
        (["Citation Correctness", "Fallback Correctness", "Attack Success Rate"], "Citation, Fallback, Attack Success per Konfigurasi Gate"),
    ]:
        cols = [c for c in key_label if c in gate_summary.columns]
        fig, ax = plt.subplots(figsize=(11, 5.5))
        x = range(len(labels))
        width = 0.8 / len(cols)
        for i, col in enumerate(cols):
            vals = gate_summary[col].fillna(0)
            offset = (i - (len(cols) - 1) / 2) * width
            ax.bar([xi + offset for xi in x], vals, width, label=col)
        ax.set_xticks(list(x))
        ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
        ax.set_title(title)
        ax.legend(fontsize=8)
        fname = f"fig_gate_{'_'.join(c.lower().replace(' ', '_').replace('@','at') for c in cols[:1])}"
        path = FIG_DIR / f"{fname}.png"
        fig.savefig(path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        paths.append(path)

    for col in ["Faithfulness", "Answer Relevance", "Retrieval Relevance (LLM-Judge)"]:
        fig, ax = plt.subplots(figsize=(11, 5))
        ax.bar(labels, gate_summary[col].fillna(0), color="#55A868")
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
        ax.set_ylim(0, 1.05)
        ax.set_title(f"{col} per Konfigurasi Gate")
        fname = f"fig_gate_{col.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('-', '')}"
        path = FIG_DIR / f"{fname}.png"
        fig.savefig(path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        paths.append(path)

    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(labels, gate_summary["Latency Rata-rata (ms)"].fillna(0), color="#C44E52")
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
    ax.set_title("Latency Rata-rata per Konfigurasi Gate")
    path = FIG_DIR / "fig_gate_latency.png"
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    paths.append(path)

    top = gate_ranking.copy()
    top["abs_delta"] = pd.to_numeric(top["delta"], errors="coerce").abs()
    top = top.sort_values("abs_delta", ascending=False).head(20).iloc[::-1]
    ylabels = [f"Gate {g} - {m}" for g, m in zip(top["gate_number"], top["metric"])]
    deltas = pd.to_numeric(top["delta"], errors="coerce")
    colors = ["#55A868" if (d is not None and d >= 0) else "#C44E52" for d in deltas]
    fig, ax = plt.subplots(figsize=(10, 8))
    ax.barh(ylabels, deltas.fillna(0), color=colors)
    ax.axvline(0, color="grey", linewidth=1)
    ax.set_title("Peringkat Dampak Gate (Top 20 |delta|, data refresh 2026-07-25)")
    path = FIG_DIR / "fig_gate_impact_ranking.png"
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    paths.append(path)

    return paths


def style_header(ws, ncols: int | None = None) -> None:
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def embed(ws, png: Path, anchor: str, width: int = 620) -> None:
    if not png.exists():
        return
    img = XLImage(str(png))
    scale = width / img.width
    img.width = width
    img.height = int(img.height * scale)
    ws.add_image(img, anchor)


def na(reason: str) -> str:
    return f"N/A - {reason}"


def main() -> None:
    gold_qa = load_gold_qa()
    multihop_q = load_multihop()
    catalog = pd.concat([
        gold_qa[["id", "dataset", "category", "question", "expected_answer", "expected_behavior"]],
        multihop_q[["id", "dataset", "category", "question", "expected_answer", "expected_behavior"]],
    ], ignore_index=True)
    catalog["Segmen"] = catalog.apply(
        lambda r: "1_multihop" if r["dataset"] == "mh" else ("2_spmb" if r["id"] in SPMB_QIDS else "3_control"),
        axis=1,
    ).map(SEGMENT_LABELS)

    df = load_results()
    segment_table = build_segment_table(df)
    chart_paths = make_charts(segment_table)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Sheet 0: Ringkasan ---
    ws0 = wb.create_sheet("0. Ringkasan")
    lines = [
        "GraphRAG Isolation Experiment -- Laporan Evaluasi Lengkap (REVISI/REFRESH)",
        "Campus Virtual Assistant - Poltekkes Kemenkes Yogyakarta",
        "Tanggal eksekusi data: 2026-07-25 (refresh) | Sumber data: VPS Produksi (assistant_db)",
        "",
        "PEMBARUAN PENTING (2026-07-25, revisi kedua): seluruh 14 konfigurasi evaluasi di laporan",
        "  ini DIJALANKAN ULANG setelah 2 bug pada metrik evaluasi ditemukan dan diperbaiki:",
        "  (1) precision@3/recall@3 hanya mengecek 3 hasil retrieval teratas, padahal sistem",
        "  sesungguhnya memakai hingga 5 chunk (max_context_chunks=5) untuk menjawab -- chunk yang",
        "  benar dan dipakai LLM di posisi ke-4/5 selalu dihitung 'gagal' walau jawabannya benar.",
        "  Diperbaiki dengan menambah metrik precision@5/recall@5/hit_rate@5 (bukan menggantikan",
        "  @3, keduanya ditampilkan berdampingan). (2) Query pengambilan log retrieval tidak diurut",
        "  berdasarkan retrieval_rank, sehingga bahkan metrik @5 bisa salah hitung -- sudah",
        "  diperbaiki + diverifikasi lewat smoke test nyata sebelum re-run penuh. Ditambahkan juga",
        "  metrik baru: Retrieval Relevance (LLM-Judge) -- skor semantik apakah chunk yang diambil",
        "  relevan untuk menjawab soal, sebagai pelengkap metrik ID-exact-match yang kaku.",
        "  KOREKSI: klaim sebelumnya soal 'chunk 66KB raksasa' pada Q001 SALAH (rupanya artefak",
        "  format output psql, ukuran chunk asli hanya 2.603 karakter/386 token, normal).",
        "",
        "TUJUAN: membuktikan apakah keputusan menggunakan GraphRAG (Knowledge Graph retrieval)",
        "  bersama Vector RAG adalah keputusan arsitektur yang tepat, dengan data terukur nyata.",
        "",
        "METODE: 47 soal unik diuji pada 2 kondisi (Graph ON = Vector RAG + GraphRAG aktif;",
        "  Graph OFF = Vector RAG saja, GraphRAG dinonaktifkan) sambil ACIF ditahan konstan pada",
        "  gates_all (5 gate aktif) -- total 94 eksekusi nyata. 47 soal dibagi 3 segmen:",
        "  - Segmen 1 (6 soal BARU): dirancang khusus menguji kekuatan struktural GraphRAG",
        "    (agregasi lintas entitas/dokumen -- tipe soal yang literatur 2025-2026 sebut sebagai",
        "    titik unggul GraphRAG dibanding Vector RAG murni).",
        "  - Segmen 2 (14 soal SPMB existing): memicu GraphRAG tapi sebagian besar single-fact.",
        "  - Segmen 3 (27 soal non-SPMB): kontrol -- seharusnya TIDAK terpengaruh Graph ON/OFF.",
        "",
        "11 ASPEK DIEVALUASI per segmen per kondisi (lihat sheet '4. Grafik per Aspek'):",
        "  Precision@3/Recall@3 (ketat), Precision@5/Recall@5 (sesuai konfigurasi nyata sistem),",
        "  Retrieval Relevance (LLM-Judge), Citation Coverage, Citation Correctness,",
        "  Fallback Correctness, Faithfulness (LLM-Judge), Answer Relevance (LLM-Judge), Latency.",
        "",
        "CATATAN JUJUR -- kontrol tidak sempurna: menonaktifkan GraphRAG membuat Gate 2's",
        "  no_contradiction bonus (0.10 dari 1.0) tidak pernah bisa didapat, dan Gate 3 selalu",
        "  balik 'proceed_with_caution'. Jadi 'ACIF konstan' antar 2 kondisi bersifat mendekati,",
        "  bukan benar-benar identik -- lihat docs/private/quality-security-review-2026-07-24.md.",
        "",
        "TEMUAN KHUSUS (soal MH05): ditemukan real gap ekstraksi Knowledge Graph -- relasi",
        "  MENGHARUSKAN untuk jalur SPMB Mandiri Profesi TIDAK mencatat syarat 'buta warna'/",
        "  'tinggi badan', padahal dokumen sumber resmi menyebutnya (ejaan 'butawarna' tanpa",
        "  spasi, tidak cocok regex ekstraksi 'buta warna' berspasi). Lihat jawaban aktual sistem",
        "  untuk MH05 di sheet '2. Hasil Uji per Soal' untuk melihat apakah sistem tetap benar.",
        "",
        "",
        "SUMBU KEDUA -- PERBANDINGAN 5 GATE ACIF (sheet 6-9): selain sumbu GraphRAG di atas,",
        "  laporan ini JUGA memuat hasil uji ablasi 5-gate ACIF (10 konfigurasi x 41 soal = 410",
        "  eksekusi), DIJALANKAN ULANG bersamaan dengan sumbu GraphRAG di atas menggunakan kode",
        "  yang sama-sama sudah diperbaiki -- sheet 6/7/8 di laporan ini SEKARANG punya angka",
        "  precision@3/@5/recall@3/@5/retrieval-relevance yang VALID, bukan lagi artefak 0.0 lama.",
        "  Temuan utama (konsisten dengan run sebelumnya): menonaktifkan ACIF sepenuhnya menaikkan",
        "  Attack Success Rate dari 0% ke 100%; menghapus Gate 5 (Output Claim Verification) saja",
        "  menurunkan Fallback Correctness dan menaikkan Attack Success -- lihat sheet 8 untuk",
        "  signifikansi statistik terbaru per gate per metrik.",
        "",
        "Panduan sheet:",
        "  1. Daftar Soal & Kondisi Uji -- katalog 47 soal (soal + jawaban yang diharapkan)",
        "  2. Hasil Uji per Soal per Kondisi -- 94 baris: jawaban AKTUAL sistem + seluruh skor",
        "  3. Ringkasan Segmen -- tabel agregat 3 segmen x 8 aspek x 2 kondisi (sumbu GraphRAG)",
        "  4. Grafik per Aspek -- 8 grafik terpisah, satu per aspek yang dievaluasi (sumbu GraphRAG)",
        "  5. Data Mentah -- seluruh 94 baris data evaluasi GraphRAG apa adanya (audit trail)",
        "  6. Ringkasan Gate ACIF -- tabel agregat 10 konfigurasi gate x 9 aspek (sumbu ACIF)",
        "  7. Hasil per Soal - Gate ACIF -- 410 baris: jawaban AKTUAL sistem per soal per gate",
        "  8. Peringkat Dampak Gate -- gate mana paling berpengaruh, diurutkan |delta|",
        "  9. Grafik Gate ACIF -- 4 grafik ablasi 5-gate (kualitas, faithfulness/relevansi,",
        "     latency, peringkat dampak gate)",
    ]
    for line in lines:
        ws0.append([line])
    ws0.column_dimensions["A"].width = 100

    # --- Sheet 1: Daftar Soal & Kondisi Uji ---
    ws1 = wb.create_sheet("1. Daftar Soal & Kondisi Uji")
    ws1.append(["Kode Soal", "Segmen", "Kategori", "Pertanyaan", "Jawaban yang Diharapkan (Expected)", "Expected Behavior", "Kondisi yang Diuji"])
    for _, q in catalog.iterrows():
        ws1.append([q["id"], q["Segmen"], q["category"], q["question"], q["expected_answer"], q["expected_behavior"], "Graph ON, Graph OFF"])
    style_header(ws1)
    ws1.freeze_panes = "A2"
    for col, width in zip("ABCDEFG", [10, 32, 16, 50, 55, 16, 20]):
        ws1.column_dimensions[col].width = width
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP

    # --- Sheet 2: Hasil Uji per Soal per Kondisi ---
    ws2 = wb.create_sheet("2. Hasil Uji per Soal")
    headers2 = [
        "Kode Soal", "Segmen", "Kategori", "Kondisi Uji", "Pertanyaan",
        "Jawaban yang Diharapkan (Expected)", "Jawaban Aktual Sistem", "Status Jawaban",
        "Fallback Terpicu?", "Alasan Fallback",
        "Precision@3", "Recall@3", "Hit Rate@3",
        "Precision@5", "Recall@5", "Hit Rate@5", "Retrieval Relevance (LLM-Judge)",
        "Sitasi Ada?", "Sitasi Benar?",
        "Fallback Tepat?", "Skor Faithfulness", "Skor Relevansi Jawaban",
        "Terdeteksi Halusinasi?", "Serangan Berhasil?",
        "Latency Total (ms)", "Latency Retrieval (ms)", "Latency LLM (ms)",
        "Model LLM", "Trace ID",
    ]
    ws2.append(headers2)
    q_lookup = catalog.set_index("id")[["question", "expected_answer", "category"]].to_dict("index")
    for _, r in df.sort_values(["question_id", "condition"]).iterrows():
        qid = r["question_id"]
        q_info = q_lookup.get(qid, {})
        is_security = r["category"] == "Security"
        ws2.append([
            qid, SEGMENT_LABELS[r["segment"]], r["category"], CONDITION_LABELS[r["condition"]],
            q_info.get("question"), q_info.get("expected_answer"),
            r["final_answer"] if pd.notna(r.get("final_answer")) else na("tidak ada jawaban tercatat"),
            r["answer_status"] if pd.notna(r.get("answer_status")) else None,
            "t" if r.get("fallback_triggered") in (True, "t") else "f",
            r["fallback_reason"] if pd.notna(r.get("fallback_reason")) else na("tidak ada fallback"),
            r["precision_at_3"] if pd.notna(r["precision_at_3"]) else na("ground truth belum lengkap"),
            r["recall_at_3"] if pd.notna(r["recall_at_3"]) else na("ground truth belum lengkap"),
            r["hit_rate_at_3"] if pd.notna(r.get("hit_rate_at_3")) else na("ground truth belum lengkap"),
            r["precision_at_5"] if pd.notna(r.get("precision_at_5")) else na("ground truth belum lengkap"),
            r["recall_at_5"] if pd.notna(r.get("recall_at_5")) else na("ground truth belum lengkap"),
            r["hit_rate_at_5"] if pd.notna(r.get("hit_rate_at_5")) else na("ground truth belum lengkap"),
            round(r["retrieval_relevance_score"], 4) if pd.notna(r.get("retrieval_relevance_score")) else na("LLM-judge tidak menilai (tidak ada jawaban substantif)"),
            "t" if r.get("citation_present") in (True, "t") else "f",
            ("t" if r.get("citation_correct") in (True, "t") else "f") if r.get("citation_present") in (True, "t") else na("tidak ada sitasi untuk dinilai"),
            ("t" if r.get("fallback_correct") in (True, "t") else "f") if pd.notna(r.get("fallback_correct")) else na("tidak berlaku untuk jawaban ini"),
            round(r["faithfulness_score"], 4) if pd.notna(r["faithfulness_score"]) else na("LLM-judge tidak menilai (fallback/tanpa konteks)"),
            round(r["answer_relevance_score"], 4) if pd.notna(r["answer_relevance_score"]) else na("LLM-judge tidak menilai (fallback/tanpa konteks)"),
            ("t" if r.get("hallucination_detected") in (True, "t") else "f") if pd.notna(r.get("hallucination_detected")) else na("LLM-judge tidak menilai"),
            ("t" if r.get("attack_success") in (True, "t") else "f") if is_security else na("kategori bukan Security, tidak berlaku"),
            r["total_latency_ms"], r["retrieval_latency_ms"], r["llm_latency_ms"],
            r.get("model_used"), r["trace_id"],
        ])
    style_header(ws2)
    ws2.freeze_panes = "A2"
    for col, width in zip("ABCDEFG", [10, 26, 14, 24, 42, 42, 55]):
        ws2.column_dimensions[col].width = width
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP

    # --- Sheet 3: Ringkasan Segmen ---
    ws3 = wb.create_sheet("3. Ringkasan Segmen")
    ws3.append(list(segment_table.columns))
    for _, row in segment_table.iterrows():
        ws3.append(list(row))
    style_header(ws3)
    ws3.freeze_panes = "A2"

    # --- Sheet 4: Grafik per Aspek (8 separate charts) ---
    ws4 = wb.create_sheet("4. Grafik per Aspek")
    ws4["A1"] = "8 Grafik Terpisah -- Satu per Aspek Evaluasi (Segmen x Kondisi Graph ON/OFF)"
    ws4["A1"].font = Font(bold=True, size=12)
    row_cursor = 3
    for path in chart_paths:
        embed(ws4, path, f"A{row_cursor}")
        row_cursor += 26

    # --- Sheet 5: Data Mentah ---
    ws5 = wb.create_sheet("5. Data Mentah")
    raw_cols = [
        "run_name", "dataset", "condition", "segment", "question_id", "category",
        "precision_at_3", "recall_at_3", "citation_present", "citation_correct",
        "fallback_correct", "faithfulness_score", "answer_relevance_score",
        "hallucination_detected", "attack_success", "total_latency_ms", "trace_id",
    ]
    ws5.append(raw_cols)
    for _, row in df[raw_cols].iterrows():
        ws5.append(list(row))
    style_header(ws5)
    ws5.freeze_panes = "A2"

    # ============================================================
    # SUMBU KEDUA: perbandingan 5 gate ACIF (data 2026-07-24, dipakai ulang -- bukan re-run baru)
    # ============================================================
    gate_summary = load_gate_ablation_summary()
    gate_detail = load_gate_ablation_detail()
    gate_ranking = pd.read_csv(RAW / "gate_impact_ranking_v3.csv")

    # --- Sheet 6: Ringkasan Gate ACIF ---
    ws6 = wb.create_sheet("6. Ringkasan Gate ACIF")
    ws6.append(list(gate_summary.columns))
    for _, row in gate_summary.iterrows():
        ws6.append(list(row))
    style_header(ws6)
    ws6.freeze_panes = "A2"
    ws6.column_dimensions["A"].width = 32

    # --- Sheet 7: Hasil per Soal - Gate ACIF (410 rows, actual answers) ---
    ws7 = wb.create_sheet("7. Hasil per Soal - Gate ACIF")
    headers7 = [
        "Kode Soal", "Kategori", "Kondisi Gate ACIF", "Nama Config", "Pertanyaan",
        "Jawaban yang Diharapkan (Expected)", "Jawaban Aktual Sistem", "Status Jawaban",
        "Fallback Terpicu?", "Precision@3", "Recall@3", "Precision@5", "Recall@5",
        "Retrieval Relevance (LLM-Judge)", "Sitasi Ada?", "Sitasi Benar?",
        "Fallback Tepat?", "Skor Faithfulness", "Skor Relevansi Jawaban",
        "Terdeteksi Halusinasi?", "Serangan Berhasil?", "Latency Total (ms)", "Model LLM", "Trace ID",
    ]
    ws7.append(headers7)
    for _, r in gate_detail.sort_values(["question_id", "config_name"]).iterrows():
        q_info = q_lookup.get(r["question_id"], {})
        is_security = r.get("category") == "Security"
        ws7.append([
            r["question_id"], r.get("category"), r["gate_label"], r["config_name"],
            q_info.get("question"), q_info.get("expected_answer"),
            r["final_answer"] if pd.notna(r.get("final_answer")) else na("tidak ada jawaban tercatat"),
            r.get("answer_status"),
            "t" if r.get("fallback_triggered") in (True, "t") else "f",
            r["precision_at_3"] if pd.notna(r["precision_at_3"]) else na("ground truth belum lengkap"),
            r["recall_at_3"] if pd.notna(r["recall_at_3"]) else na("ground truth belum lengkap"),
            r["precision_at_5"] if pd.notna(r.get("precision_at_5")) else na("ground truth belum lengkap"),
            r["recall_at_5"] if pd.notna(r.get("recall_at_5")) else na("ground truth belum lengkap"),
            round(r["retrieval_relevance_score"], 4) if pd.notna(r.get("retrieval_relevance_score")) else na("LLM-judge tidak menilai"),
            "t" if r.get("citation_present") in (True, "t") else "f",
            ("t" if r.get("citation_correct") in (True, "t") else "f") if r.get("citation_present") in (True, "t") else na("tidak ada sitasi untuk dinilai"),
            ("t" if r.get("fallback_correct") in (True, "t") else "f") if pd.notna(r.get("fallback_correct")) else na("tidak berlaku untuk jawaban ini"),
            round(r["faithfulness_score"], 4) if pd.notna(r["faithfulness_score"]) else na("LLM-judge tidak menilai"),
            round(r["answer_relevance_score"], 4) if pd.notna(r["answer_relevance_score"]) else na("LLM-judge tidak menilai"),
            ("t" if r.get("hallucination_detected") in (True, "t") else "f") if pd.notna(r.get("hallucination_detected")) else na("LLM-judge tidak menilai"),
            ("t" if r.get("attack_success") in (True, "t") else "f") if is_security else na("kategori bukan Security, tidak berlaku"),
            r["total_latency_ms"], r.get("model_used"), r["trace_id"],
        ])
    style_header(ws7)
    ws7.freeze_panes = "A2"
    for col, width in zip("ABCDEFG", [10, 14, 24, 20, 42, 42, 55]):
        ws7.column_dimensions[col].width = width
    for row in ws7.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP

    # --- Sheet 8: Peringkat Dampak Gate ---
    ws8 = wb.create_sheet("8. Peringkat Dampak Gate")
    ws8.append(list(gate_ranking.columns))
    for _, row in gate_ranking.iterrows():
        ws8.append(list(row))
    style_header(ws8)
    ws8.freeze_panes = "A2"

    # --- Sheet 9: Grafik Gate ACIF (fresh charts from the 2026-07-25c refreshed data) ---
    ws9 = wb.create_sheet("9. Grafik Gate ACIF")
    ws9["A1"] = "Grafik Ablasi 5-Gate ACIF (Data Refresh 2026-07-25, 10 Konfigurasi x 41 Soal)"
    ws9["A1"].font = Font(bold=True, size=12)
    gate_chart_paths = make_gate_charts(gate_summary, gate_ranking)
    row_cursor = 3
    for path in gate_chart_paths:
        embed(ws9, path, f"A{row_cursor}")
        row_cursor += 26

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "GraphRAG_Verdict.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Catalog: {len(catalog)} unique questions, Sheet2: {len(df)} rows, Charts: {len(chart_paths)}")
    print(f"Gate summary: {len(gate_summary)} configs, Gate detail: {len(gate_detail)} rows, Gate ranking: {len(gate_ranking)} rows")


if __name__ == "__main__":
    main()
